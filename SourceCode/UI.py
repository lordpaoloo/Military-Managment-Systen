from PyQt5.QtCore import QTimer, Qt
from PyQt5 import QtCore, QtWidgets, QtGui
import sqlite3
import openpyxl
from openpyxl.utils import get_column_letter
from PyQt5.QtPrintSupport import QPrintDialog, QPrinter, QPrinterInfo, QPrintPreviewDialog
from PyQt5.QtWidgets import (
    QComboBox,
    QApplication,
    QDialog,
    QVBoxLayout,
    QLabel,
    QLineEdit,
    QPushButton,
    QMessageBox,
    QFileDialog,
    QHBoxLayout,
)
from hijri_converter import convert
from datetime import datetime, timedelta, date
from PyQt5.QtGui import QIcon, QPainter, QPixmap, QImage, QPdfWriter
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4, A3
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.pdfbase import pdfmetrics
from reportlab.lib.utils import ImageReader
from reportlab.lib import colors
from reportlab.platypus import Table, TableStyle
from bidi.algorithm import get_display
import arabic_reshaper
from typing import List
import aspose.pdf as ap
import fitz  # PyMuPDF
import sys


##############################  UI Dialog ################################
class report_bydate(QDialog):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("أدخل فترة التاريخ")

        # التخطيط
        layout = QVBoxLayout()

        # التسمية والإدخال
        self.label_from = QLabel("أدخل تاريخ البداية (YYYY-MM-DD):")
        self.date_from_input = QLineEdit(self)
        self.date_from_input.setPlaceholderText("1446-03-05")

        self.label_to = QLabel("أدخل تاريخ النهاية (YYYY-MM-DD):")
        self.date_to_input = QLineEdit(self)
        self.date_to_input.setPlaceholderText("1446-03-05")

        # زر الإرسال
        self.submit_button = QPushButton("إرسال")
        self.submit_button.clicked.connect(self.on_submit)

        # إضافة العناصر إلى التخطيط
        layout.addWidget(self.label_from)
        layout.addWidget(self.date_from_input)
        layout.addWidget(self.label_to)
        layout.addWidget(self.date_to_input)
        layout.addWidget(self.submit_button)

        self.setLayout(layout)

    def on_submit(self):
        date_from = self.date_from_input.text()
        date_to = self.date_to_input.text()

        # التحقق من صحة الإدخال
        if not self.validate_date(date_from) or not self.validate_date(date_to):
            QMessageBox.warning(self, "إدخال غير صالح", "يرجى إدخال تواريخ صالحة بالصيغة YYYY-MM-DD.")
            return

        # استعلام قاعدة البيانات
        data = self.query_database(date_from, date_to)

        if data:
                        # طلب تحديد مكان حفظ ملف PDF
            output_filename, _ = QFileDialog.getSaveFileName(
                self, "حفظ ملف Excel", "", "ملفات Excel (*.xlsx)"
            )

            if output_filename:
                self.create_excel(output_filename, data, date_from, date_to)
                QMessageBox.information(self, "نجاح", "تم إنشاء ملف PDF بنجاح.")
        else:
            QMessageBox.information(
                self, "لا توجد بيانات", "لم يتم العثور على سجلات للفترة المدخلة."
            )

    def validate_date(self, date_text):
        # التحقق من الصيغة YYYY-MM-DD
        try:
            datetime.strptime(date_text, "%Y-%m-%d")
            return True
        except ValueError:
            return False

    def query_database(self, date_from, date_to):
        conn = sqlite3.connect("agents.db")
        cursor = conn.cursor()

        # Query to match the start_Date within the range
        query = """
        SELECT military_number, the_kind_of_holiday, duration_of_vacation, start_Date, return_date, duration_of_absence, user_code, check_in_date
        FROM holidays_history
        WHERE start_Date BETWEEN ? AND ?
        """

        cursor.execute(query, (date_from, date_to))
        rows = cursor.fetchall()
        conn.close()

        # Format data
        formatted_data = []
        for row in rows:
            name = self.get_name_by_military_number(
                row[0]
            )  # Retrieve name using military number
            user_name = self.get_username_by_code(
                row[6]
            )  # Retrieve user_name using user_code
            formatted_row = [
                row[0],  # military_number
                name,  # name
                row[1],  # the_kind_of_holiday
                row[2],  # duration_of_vacation
                row[3],  # start_Date (Hijri)
                row[4],  # return_date (Hijri)
                row[5],  # duration_of_absence
                user_name,  # user_name
                row[7],  # check_in_date
            ]
            formatted_data.append(formatted_row)

        return formatted_data

    def get_name_by_military_number(self, military_number):
        conn = sqlite3.connect("agents.db")
        cursor = conn.cursor()
        cursor.execute(
            "SELECT name FROM users WHERE military_number = ?", (military_number,)
        )
        result = cursor.fetchone()
        conn.close()
        return result[0] if result else "غير معروف"

    def get_username_by_code(self, user_code):
        conn = sqlite3.connect("agents.db")
        cursor = conn.cursor()
        cursor.execute(
            "SELECT user_name FROM admin_users WHERE user_code = ?", (user_code,)
        )
        result = cursor.fetchone()
        conn.close()
        return result[0] if result else "غير معروف"
    def create_excel(self, output_filename, table_data, date_from, date_to):
        # Create a new Excel workbook and select the active worksheet
        wb = openpyxl.Workbook()
        ws = wb.active

        # Set the sheet name
        ws.title = "Holiday Data"

        # Add the report title with the date range at the top
        ws.merge_cells('A1:I1')
        ws['A1'] = f"تقرير من {date_from} إلى {date_to}"
        ws['A1'].font = openpyxl.styles.Font(size=14, bold=True)
        ws['A1'].alignment = openpyxl.styles.Alignment(horizontal='center')

        # Define headers in Arabic (corresponding to your table headers)
        headers = [
            "اسم المستخدم",
            "مدة التطويف",
            "تاريخ العودة",
            "تاريخ النهاية",
            "تاريخ البداية",
            "مدة الاجازة",
            "نوع الاجازة",
            "الاسم",
            "الرقم الخاص",
        ]

        # Add headers to the second row
        for col_num, header in enumerate(headers, start=1):
            col_letter = get_column_letter(col_num)
            ws[f'{col_letter}2'] = header
            ws[f'{col_letter}2'].font = openpyxl.styles.Font(bold=True)
            ws[f'{col_letter}2'].alignment = openpyxl.styles.Alignment(horizontal='center')

        # Reorder the data to match headers
        for row_num, row_data in enumerate(table_data, start=3):
            sorted_row = [
                row_data[7],  # اسم المستخدم
                row_data[6],  # مدة التطويف
                row_data[8],  # تاريخ العودة
                row_data[5],  # تاريخ النهاية
                row_data[4],  # تاريخ البداية
                row_data[3],  # مدة الاجازة
                row_data[2],  # نوع الاجازة
                row_data[1],  # الاسم
                row_data[0],  # الرقم الخاص
            ]

            # Add sorted row to Excel sheet
            for col_num, cell_value in enumerate(sorted_row, start=1):
                col_letter = get_column_letter(col_num)
                ws[f'{col_letter}{row_num}'] = cell_value
                ws[f'{col_letter}{row_num}'].alignment = openpyxl.styles.Alignment(horizontal='center')

        # Adjust column widths
        for col_num in range(1, len(headers) + 1):
            col_letter = get_column_letter(col_num)
            ws.column_dimensions[col_letter].width = 20  # Adjust width as needed

        # Add footer with report generation info
        last_row = len(table_data) + 4  # Adjust row for footer (after table)
        ws.merge_cells(f'A{last_row}:I{last_row}')
        ws[f'A{last_row}'] = f"تقرير تم إنشاؤه في الفترة: {date_from} - {date_to}"
        ws[f'A{last_row}'].font = openpyxl.styles.Font(italic=True)
        ws[f'A{last_row}'].alignment = openpyxl.styles.Alignment(horizontal='center')

        # Save the Excel file
        wb.save(output_filename)
class report(QDialog):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("أدخل السنة الهجرية")

        # التخطيط
        layout = QVBoxLayout()

        # التسمية والإدخال
        self.label = QLabel("أدخل السنة الهجرية:")
        self.year_input = QLineEdit(self)

        # زر الإرسال
        self.submit_button = QPushButton("إرسال")
        self.submit_button.clicked.connect(self.on_submit)

        # إضافة العناصر إلى التخطيط
        layout.addWidget(self.label)
        layout.addWidget(self.year_input)
        layout.addWidget(self.submit_button)

        self.setLayout(layout)

    def on_submit(self):
        hijri_year = self.year_input.text()

        # التحقق من صحة الإدخال
        if not hijri_year.isdigit():
            QMessageBox.warning(self, "إدخال غير صالح", "يرجى إدخال سنة هجرية صالحة.")
            return

        # استعلام قاعدة البيانات
        data = self.query_database(hijri_year)

        if data:
            # طلب تحديد مكان حفظ ملف PDF
            output_filename, _ = QFileDialog.getSaveFileName(
                self, "حفظ PDF", "", "ملفات PDF (*.pdf)"
            )

            if output_filename:
                self.create_pdf(output_filename, data, hijri_year)
                QMessageBox.information(self, "نجاح", "تم إنشاء ملف PDF بنجاح.")
        else:
            QMessageBox.information(
                self, "لا توجد بيانات", "لم يتم العثور على سجلات للسنة المدخلة."
            )

    def query_database(self, hijri_year):
        conn = sqlite3.connect("agents.db")
        cursor = conn.cursor()

        # Query to match only the year in the start_Date column
        query = """
        SELECT military_number, the_kind_of_holiday, duration_of_vacation, start_Date, return_date, duration_of_absence, user_code, check_in_date
        FROM holidays_history
        WHERE SUBSTR(start_Date, 1, 4) = ?
        """

        cursor.execute(query, (hijri_year,))
        rows = cursor.fetchall()
        conn.close()

        # Format data
        formatted_data = []
        for row in rows:
            name = self.get_name_by_military_number(
                row[0]
            )  # Retrieve name using military number
            user_name = self.get_username_by_code(
                row[6]
            )  # Retrieve user_name using user_code
            formatted_row = [
                row[0],  # military_number
                name,  # name
                row[1],  # the_kind_of_holiday
                row[2],  # duration_of_vacation
                row[3],  # start_Date (Hijri)
                row[4],  # return_date (Hijri)
                row[5],  # duration_of_absence
                user_name,  # user_name
                row[7],  # check_in_date
            ]
            formatted_data.append(formatted_row)

        return formatted_data

    def get_name_by_military_number(self, military_number):
        conn = sqlite3.connect("agents.db")
        cursor = conn.cursor()
        cursor.execute(
            "SELECT name FROM users WHERE military_number = ?", (military_number,)
        )
        result = cursor.fetchone()
        conn.close()
        return result[0] if result else "غير معروف"

    def get_username_by_code(self, user_code):
        conn = sqlite3.connect("agents.db")
        cursor = conn.cursor()
        cursor.execute(
            "SELECT user_name FROM admin_users WHERE user_code = ?", (user_code,)
        )
        result = cursor.fetchone()
        conn.close()
        return result[0] if result else "غير معروف"

    def create_pdf(self, output_filename, table_data, year):
        c = canvas.Canvas(output_filename, pagesize=A3)  # Set page size to A3
        width, height = A3

        # Register the custom Arabic font
        pdfmetrics.registerFont(TTFont("Arabic", "Printing/Amiri-Regular.ttf"))

        # Load the image
        image_path = f"Printing/logo.png"  # Replace with the path to your image
        image = ImageReader(image_path)
        image_width, image_height = image.getSize()
        aspect_ratio = image_height / float(image_width)
        image_display_width = width / 8  # Adjust as needed
        image_display_height = image_display_width * aspect_ratio
        image_x = (width - image_display_width) / 2
        image_y = height - image_display_height - 30

        # Draw the image
        c.drawImage(
            image_path,
            image_x,
            image_y,
            width=image_display_width,
            height=image_display_height,
        )

        # Arabic text
        arabic_text1 = "المملكة العربية السعودية"
        reshaped_text1 = arabic_reshaper.reshape(arabic_text1)
        bidi_text1 = get_display(reshaped_text1)

        arabic_text2 = "رئاسة أمن الدولة"
        reshaped_text2 = arabic_reshaper.reshape(arabic_text2)
        bidi_text2 = get_display(reshaped_text2)

        arabic_text3 = "قوات الطوارئ الخاصة"
        reshaped_text3 = arabic_reshaper.reshape(arabic_text3)
        bidi_text3 = get_display(reshaped_text3)

        # Add text to PDF
        text_y = 1100  # Adjusted for A3 size
        c.setFont("Arabic", 10)
        c.drawRightString(width - 20, text_y, bidi_text1)
        c.drawRightString(width - 20, text_y - 20, bidi_text2)
        c.drawRightString(width - 20, text_y - 40, bidi_text3)

        # Table headers
        headers = [
            "اسم المستخدم",
            "مدة التطويف",
            "تاريخ العودة",
            "تاريخ النهاية",
            "تاريخ البداية",
            "مدة الاجازة",
            "نوع الاجازة",
            "الاسم",
            "الرقم الخاص",
        ]

        # Reorder the data to match headers
        sorted_data = [headers]
        for row in table_data:
            sorted_row = [
                row[7],  # الرقم الخاص
                row[6],  # الاسم
                row[8],  # نوع الاجازة
                row[5],  # مدة الاجازة
                row[4],  # تاريخ البدابة
                row[3],  # تاريخ النهاية
                row[2],  # تاريخ العودة
                row[1],  # مدة التطويف
                row[0],  # اسم المستخدم
            ]
            sorted_data.append(sorted_row)

        # Reshape Arabic text in the table
        reshaped_data = []
        for row in sorted_data:
            reshaped_row = []
            for item in row:
                if isinstance(item, str):
                    reshaped_item = arabic_reshaper.reshape(item)
                    bidi_item = get_display(reshaped_item)
                    reshaped_row.append(bidi_item)
                else:
                    reshaped_row.append(item)
            reshaped_data.append(reshaped_row)

        table = Table(reshaped_data)
        table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                    ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                    ("FONTNAME", (0, 0), (-1, -1), "Arabic"),
                    ("FONTSIZE", (0, 0), (-1, -1), 10),
                    ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
                    ("BACKGROUND", (0, 1), (-1, -1), colors.beige),
                    ("GRID", (0, 0), (-1, -1), 1, colors.black),
                ]
            )
        )

        table.wrapOn(c, width, height)
        table.drawOn(c, 50, text_y - 250)  # Adjust for larger page

        # Save PDF
        c.save()
class UpdateAdmin(QDialog):
    def __init__(self):
        super().__init__()

        self.setWindowTitle("تغير الادمن و الكود")
        self.setGeometry(300, 300, 400, 200)

        # Create layout
        layout = QVBoxLayout()

        # Create username label and input
        self.username_label = QLabel("الاسم:")
        self.username_input = QLineEdit()
        layout.addWidget(self.username_label)
        layout.addWidget(self.username_input)

        # Create password label and input
        self.password_label = QLabel("الكود:")
        self.password_input = QLineEdit()
        self.password_input.setEchoMode(QLineEdit.Password)  # Hide password input
        layout.addWidget(self.password_label)
        layout.addWidget(self.password_input)

        # Create submit button
        self.submit_button = QPushButton("تغيبر")
        self.submit_button.clicked.connect(self.update_admin_credentials)
        layout.addWidget(self.submit_button)

        self.setLayout(layout)

    def update_admin_credentials(self):
        new_username = self.username_input.text()
        new_password = self.password_input.text()

        if new_username and new_password:
            try:
                # Connect to the SQLite database
                conn = sqlite3.connect("agents.db")
                cursor = conn.cursor()

                # Update the first row's username and password in the admin table
                cursor.execute(
                    """
                    UPDATE admin
                    SET username = ?, password = ?
                    WHERE ROWID = (SELECT ROWID FROM admin LIMIT 1)
                """,
                    (new_username, int(new_password)),
                )
                cursor.execute(
                    """
                    UPDATE admin_users
                    SET user_name = ?, user_code = ?
                    WHERE ROWID = (SELECT ROWID FROM admin LIMIT 1)
                """,
                    (new_username, int(new_password)),
                )
                # Commit the changes
                conn.commit()
                # Show success message
                QMessageBox.information(
                    self, "Success", "Admin credentials updated successfully."
                )

            except Exception as e:
                QMessageBox.critical(self, "Error", f"An error occurred: {e}")

            finally:
                # Close the connection
                conn.close()

                # Close the dialog
                self.accept()
        else:
            QMessageBox.warning(
                self, "Input Error", "Please enter both username and password."
            )
class PrintDialog(QDialog):
    def __init__(self, db_path, militarycode, name):
        super().__init__()
        self.db_path = db_path
        self.militarycode = militarycode
        self.name = name
        self.setWindowTitle("Holiday History")
        self.setGeometry(100, 100, 800, 600)

        self.layout = QVBoxLayout()

        # Date input layout
        date_layout = QHBoxLayout()

        self.start_date_label = QLabel("من التاريخ:")
        self.start_date_edit = QLineEdit()
        today_gregorian = datetime.today()

        # Convert to Hijri date
        today_hijri = convert.Gregorian(today_gregorian.year, today_gregorian.month, today_gregorian.day).to_hijri()
        self.start_date_edit.setText(str(today_hijri))

        self.end_date_label = QLabel("الي التاريخ:")
        self.end_date_edit = QLineEdit()
        self.end_date_edit.setText(str(today_hijri))

        date_layout.addWidget(self.end_date_edit)
        date_layout.addWidget(self.end_date_label)
        date_layout.addWidget(self.start_date_edit)
        date_layout.addWidget(self.start_date_label)
        self.layout.addLayout(date_layout)

        # Printer selection layout


        # Print button
        self.print_button = QPushButton("Print")
        self.print_button.clicked.connect(self.handle_print_button)

        self.layout.addWidget(self.print_button)
        self.setLayout(self.layout)
    def print_pdf_file(self, file_path):
        if not file_path:
            QMessageBox.critical(self, "Error", "No file selected.")
            return

        # Create a printer object
        printer = QPrinter(QPrinter.HighResolution)

        # Open the PDF document
        pdf_document = fitz.open(file_path)
        if pdf_document.page_count > 0:
            # Show the print dialog to choose a printer
            print_dialog = QPrintDialog(printer, self)
            if print_dialog.exec_() == QPrintDialog.Accepted:
                # Start printing the document
                painter = QPainter(printer)

                for page_number in range(pdf_document.page_count):
                    if page_number > 0:
                        printer.newPage()

                    image = self.render_page(pdf_document, page_number, printer.pageRect().size())
                    painter.drawImage(0, 0, image)

                painter.end()
            else:
                print("Print dialog was canceled.")
        else:
            QMessageBox.critical(self, "Error", "Failed to load PDF document.")

    def render_page(self, pdf_document, page_number, size):
        page = pdf_document.load_page(page_number)  # Load a single page
        # Get the size of the page in the PDF
        pdf_width, pdf_height = page.rect.width, page.rect.height
        
        # Calculate the scaling factor to fit the PDF page to the printer page size
        scale_x = size.width() / pdf_width
        scale_y = size.height() / pdf_height
        scale = min(scale_x, scale_y)  # Use the smaller scale to maintain aspect ratio
        
        # Render the page with the calculated scale
        pix = page.get_pixmap(matrix=fitz.Matrix(scale, scale))
        return QImage(pix.samples, pix.width, pix.height, pix.stride, QImage.Format_RGB888)
    def handle_print_button(self):
        start_date_hijri = self.start_date_edit.text()
        end_date_hijri = self.end_date_edit.text()

        self.print(start_date_hijri, end_date_hijri, self.militarycode)

    def print(self, start_date_hijri: str, end_date_hijri: str, military_code: int = None):
        table = self.get_holiday_data(start_date_hijri, end_date_hijri, military_code)
        self.create_pdf("Printing/holidays_history.pdf", table, self.name, self.get_civil_registry_by_military_number(military_code))
        
        self.print_pdf_file("Printing/holidays_history.pdf")

    def get_civil_registry_by_military_number(self, military_number):
        conn = sqlite3.connect("agents.db")
        cursor = conn.cursor()
        query = "SELECT civil_registry FROM users WHERE military_number = ?"
        cursor.execute(query, (military_number,))
        result = cursor.fetchone()
        conn.close()

        if result:
            return result[0]
        else:
            return "Name not found"

    def get_holiday_data(self, start_date_hijri: str, end_date_hijri: str, military_code: int = None) -> List[List]:
        conn = sqlite3.connect("agents.db")
        cursor = conn.cursor()

        query = """
        SELECT military_number, the_kind_of_holiday, duration_of_vacation, start_Date, return_date, duration_of_absence, user_code, check_in_date
        FROM holidays_history
        WHERE start_Date BETWEEN ? AND ?
        """

        if military_code is not None:
            query += " AND military_number = ?"
            params = (start_date_hijri, end_date_hijri, military_code)
        else:
            params = (start_date_hijri, end_date_hijri)

        cursor.execute(query, params)
        rows = cursor.fetchall()
        conn.close()

        formatted_data = []
        for row in rows:
            formatted_row = [
                row[0],
                row[1],
                row[2],
                row[3],
                row[4],
                row[5],
                row[6],
                row[7],
            ]
            formatted_data.append(formatted_row)

        return formatted_data

    def show_printer_check_dialog(self):
        msg_box = QMessageBox()
        msg_box.setIcon(QMessageBox.Warning)
        msg_box.setWindowTitle("تحذير")
        msg_box.setText("الرجاء التحقق من الطابعه")
        msg_box.setStandardButtons(QMessageBox.Ok)
        msg_box.exec_()

    def printer_check_done(self):
        msg_box = QMessageBox()
        msg_box.setIcon(QMessageBox.Warning)
        msg_box.setWindowTitle("تحذير")
        msg_box.setText("تم ارسال الملف الي الطابعه")
        msg_box.setStandardButtons(QMessageBox.Ok)
        msg_box.exec_()
    def create_pdf(self, output_filename, table_data, user_name, user_code):
        c = canvas.Canvas(output_filename, pagesize=A4)
        width, height = A4

        # Register the custom Arabic font
        pdfmetrics.registerFont(TTFont("Arabic", "Printing/Amiri-Regular.ttf"))

        # Load the image
        image_path = "Printing/logo.png"  # Replace with the path to your image
        image = ImageReader(image_path)
        image_width, image_height = image.getSize()
        aspect_ratio = image_height / float(image_width)
        image_display_width = width / 8  # Adjust as needed
        image_display_height = image_display_width * aspect_ratio
        image_x = (width - image_display_width) / 2
        image_y = height - image_display_height - 30

        # Draw the image
        c.drawImage(
            image_path,
            image_x,
            image_y,
            width=image_display_width,
            height=image_display_height,
        )

        # Arabic text
        arabic_text1 = "المملكة العربية السعودية"
        reshaped_text1 = arabic_reshaper.reshape(arabic_text1)
        bidi_text1 = get_display(reshaped_text1)

        arabic_text2 = "رئاسة أمن الدولة"
        reshaped_text2 = arabic_reshaper.reshape(arabic_text2)
        bidi_text2 = get_display(reshaped_text2)

        arabic_text3 = "قوات الطوارئ الخاصة"
        reshaped_text3 = arabic_reshaper.reshape(arabic_text3)
        bidi_text3 = get_display(reshaped_text3)

        # Add text to PDF
        text_y = 800
        c.setFont("Arabic", 10)
        c.drawRightString(width - 20, text_y, bidi_text1)
        c.drawRightString(width - 20, text_y - 20, bidi_text2)
        c.drawRightString(width - 20, text_y - 40, bidi_text3)

        # Additional text on the other side of the page
        arabic_text4 = f" رقم الهوية : {user_code}"
        reshaped_text4 = arabic_reshaper.reshape(arabic_text4)
        bidi_text4 = get_display(reshaped_text4)

        arabic_text5 = f"الاسم : {user_name}"
        reshaped_text5 = arabic_reshaper.reshape(arabic_text5)
        bidi_text5 = get_display(reshaped_text5)

        c.drawRightString(150, text_y, bidi_text4)
        c.drawRightString(150, text_y - 20, bidi_text5)

        # Table headers
        headers = [
            "اسم المستخدم",
            "مدة التطويف",
            "تاريخ العودة",
            "تاريخ النهاية",
            "تاريخ البدابة",
            "مدة الاجازة",
            "نوع الاجازة",
            "الرقم الخاص",
        ]

        # Reorder the data to match headers
        sorted_data = [headers]
        for row in table_data:
            sorted_row = [
                row[6],  # الرقم الخاص
                row[5],  # نوع الاجازة
                row[7],  # مدة الاجازة
                row[4],  # تاريخ البدابة
                row[3],  # تاريخ النهاية
                row[2],  # تاريخ العودة
                row[1],  # مدة التطويف
                row[0],  # اسم المستخدم
            ]
            sorted_data.append(sorted_row)

        # Reshape Arabic text in the table
        reshaped_data = []
        for row in sorted_data:
            reshaped_row = []
            for item in row:
                if isinstance(item, str):
                    reshaped_item = arabic_reshaper.reshape(item)
                    bidi_item = get_display(reshaped_item)
                    reshaped_row.append(bidi_item)
                else:
                    reshaped_row.append(item)
            reshaped_data.append(reshaped_row)

        # Create the table
        table = Table(reshaped_data)
        table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                    ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                    ("FONTNAME", (0, 0), (-1, -1), "Arabic"),
                    ("FONTSIZE", (0, 0), (-1, -1), 10),
                    ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
                    ("BACKGROUND", (0, 1), (-1, -1), colors.beige),
                    ("GRID", (0, 0), (-1, -1), 1, colors.black),
                ]
            )
        )

        # Automatically split table across multiple pages
        available_height = text_y - 150  # Adjust as needed to fit your layout
        table.wrapOn(c, width, available_height)
        
        # Draw table on multiple pages if necessary
        table_height = table._height
        if table_height > available_height:
            table_parts = table.split(width, available_height)
            for part in table_parts:
                part.wrapOn(c, width, available_height)
                part.drawOn(c, 50, available_height - part._height)
                c.showPage()  # Create new page
                available_height = height - 100  # Reset available height for new page
        else:
            table.drawOn(c, 50, text_y - 150)

        # Save PDF
        c.save()
class EditHistory(QtWidgets.QDialog):
        def __init__(self, military_number):
            super().__init__()

            self.military_number = military_number
            self.initUI()
            self.populate_holidays_table()

        def initUI(self):
            self.setObjectName("EditHistoryDialog")
            self.resize(1000, 800)  # Larger size for the dialog

            # Set up grid layout
            self.gridLayout = QtWidgets.QGridLayout(self)
            self.gridLayout.setObjectName("gridLayout")

            # Create and set up holidays table
            self.Holidays_History = QtWidgets.QTableWidget(self)
            self.Holidays_History.setObjectName("Holidays_History")
            self.Holidays_History.setLayoutDirection(QtCore.Qt.RightToLeft)  # Right-to-Left layout
            self.Holidays_History.setColumnCount(8)
            self.Holidays_History.setHorizontalHeaderLabels(
                [
                    "الرقم الخاص",
                    "نوع الاجازة",
                    "مدة الاجازة",
                    "تاربخ البداية",
                    "تاربخ النهاية",
                    "مدة التطويف",
                    "اسم المستخدم",
                    "تاريخ العودة",
                ]
            )
            self.gridLayout.addWidget(self.Holidays_History, 0, 0, 1, 1)

            # Create and set up buttons
            self.delete_button = QtWidgets.QPushButton(self)
            self.delete_button.setObjectName("delete_button")
            self.delete_button.setText("حذف")
            self.gridLayout.addWidget(self.delete_button, 1, 0, 1, 1)
            # Create and set up dialog button box
            self.buttonBox = QtWidgets.QDialogButtonBox(self)
            self.buttonBox.setOrientation(QtCore.Qt.Horizontal)
            self.buttonBox.setStandardButtons(QtWidgets.QDialogButtonBox.Cancel | QtWidgets.QDialogButtonBox.Ok)
            self.buttonBox.setObjectName("buttonBox")
            self.gridLayout.addWidget(self.buttonBox, 4, 0, 1, 1)

            # Connect signals and slots
            self.buttonBox.accepted.connect(self.accept)
            self.buttonBox.rejected.connect(self.reject)
            self.delete_button.clicked.connect(self.delete_selected_row)

            QtCore.QMetaObject.connectSlotsByName(self)

        def get_holidays_data(self):
            connection = sqlite3.connect("agents.db")
            cursor = connection.cursor()
            cursor.execute(
                """
                SELECT military_number, the_kind_of_holiday, duration_of_vacation, start_Date, return_date, duration_of_absence, user_code, check_in_date
                FROM holidays_history
                WHERE military_number = ?
            """,
                (self.military_number,),
            )
            data = cursor.fetchall()
            connection.close()
            return data

        def populate_holidays_table(self):
            data = self.get_holidays_data()
            self.Holidays_History.setRowCount(len(data))

            for row_num, row_data in enumerate(data):
                for col_num, col_data in enumerate(row_data):
                    item = QtWidgets.QTableWidgetItem(str(col_data))
                    if col_num == 0:  # Military Number column
                        item.setFlags(
                            item.flags() & ~QtCore.Qt.ItemIsEditable
                        )  # Make read-only
                    self.Holidays_History.setItem(row_num, col_num, item)

        def save_changes(self):
            connection = sqlite3.connect("agents.db")
            cursor = connection.cursor()

            row_count = self.Holidays_History.rowCount()
            col_count = self.Holidays_History.columnCount()

            for row in range(row_count):
                row_data = []
                for column in range(col_count):
                    item = self.Holidays_History.item(row, column)
                    row_data.append(item.text() if item else "")

                cursor.execute(
                    """
                    UPDATE holidays_history
                    SET the_kind_of_holiday=?, duration_of_vacation=?, start_Date=?, return_date=?, duration_of_absence=?, user_code=?, check_in_date=?
                    WHERE military_number=?
                """,
                    (
                        row_data[1],
                        row_data[2],
                        row_data[3],
                        row_data[4],
                        row_data[5],
                        row_data[6],
                        row_data[7],
                        row_data[0],
                    ),
                )

            connection.commit()
            connection.close()

            # Reset the background color of all cells
            for row in range(row_count):
                for column in range(col_count):
                    item = self.Holidays_History.item(row, column)
                    item.setBackground(QtGui.QColor("white"))

        def delete_selected_row(self):
            selected_row = self.Holidays_History.currentRow()
            if selected_row >= 0:
                # Get the military number and holiday start date of the selected row
                military_number = self.Holidays_History.item(selected_row, 0).text()
                start_date = self.Holidays_History.item(selected_row, 3).text()

                # Remove the row from the UI
                self.Holidays_History.removeRow(selected_row)

                # Delete the row from the database
                connection = sqlite3.connect("agents.db")
                cursor = connection.cursor()
                cursor.execute(
                    """
                    DELETE FROM holidays_history
                    WHERE military_number = ? AND start_Date = ?
                    """,
                    (military_number, start_date)
                )
                connection.commit()
                connection.close()
            msg_box = QtWidgets.QMessageBox()
            msg_box.setIcon(QtWidgets.QMessageBox.Warning)
            msg_box.setWindowTitle(" انتبه")
            msg_box.setText("الرجاء تعديل المتبقي من الاجازات لهذا المستخدم ليتمشي مع اي تعديل تم تغيره")
            msg_box.setStandardButtons(QtWidgets.QMessageBox.Ok)
            msg_box.exec_()
class AdminUsersDialog(QtWidgets.QDialog):  
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setObjectName("AdminUsersDialog")
        self.resize(263, 401)  # Adjusted size

        # Grid layout for the dialog
        self.gridLayout = QtWidgets.QGridLayout(self)
        self.gridLayout.setObjectName("gridLayout")

        # QTableWidget setup
        self.admin_users = QtWidgets.QTableWidget(self)
        self.admin_users.setObjectName("admin_users")
        self.admin_users.setColumnCount(2)
        self.admin_users.setHorizontalHeaderLabels(["الاسم", "الكود"])
        self.gridLayout.addWidget(self.admin_users, 0, 0, 1, 1)

        # Form layout for inputs
        self.formLayout = QtWidgets.QFormLayout()
        self.formLayout.setObjectName("formLayout")
        self.user_name_input = QtWidgets.QLineEdit(self)
        self.user_name_input.setObjectName("user_name_input")
        self.formLayout.setWidget(3, QtWidgets.QFormLayout.FieldRole, self.user_name_input)
        self.code = QtWidgets.QLabel(self)
        self.code.setObjectName("code")
        self.formLayout.setWidget(4, QtWidgets.QFormLayout.FieldRole, self.code)
        self.code_input = QtWidgets.QLineEdit(self)
        self.code_input.setObjectName("code_input")
        self.formLayout.setWidget(5, QtWidgets.QFormLayout.FieldRole, self.code_input)
        self.user_name = QtWidgets.QLabel(self)
        self.user_name.setObjectName("user_name")
        self.formLayout.setWidget(2, QtWidgets.QFormLayout.FieldRole, self.user_name)

        # Add the "Add", "Edit", and "Save" buttons
        self.add_user_btn = QtWidgets.QPushButton(self)
        self.add_user_btn.setObjectName("add_user_btn")
        self.formLayout.setWidget(6, QtWidgets.QFormLayout.FieldRole, self.add_user_btn)

        self.edit_user_btn = QtWidgets.QPushButton(self)
        self.edit_user_btn.setObjectName("edit_user_btn")
        self.formLayout.setWidget(7, QtWidgets.QFormLayout.FieldRole, self.edit_user_btn)
        self.edit_user_btn.setEnabled(False)  # Disable the edit button initially

        self.save_user_btn = QtWidgets.QPushButton(self)
        self.save_user_btn.setObjectName("save_user_btn")
        self.formLayout.setWidget(8, QtWidgets.QFormLayout.FieldRole, self.save_user_btn)
        self.save_user_btn.setEnabled(False)  # Disable the save button initially

        self.gridLayout.addLayout(self.formLayout, 1, 0, 1, 1)

        # ButtonBox for dialog buttons
        self.buttonBox = QtWidgets.QDialogButtonBox(self)
        self.buttonBox.setOrientation(QtCore.Qt.Horizontal)
        self.buttonBox.setStandardButtons(QtWidgets.QDialogButtonBox.Cancel|QtWidgets.QDialogButtonBox.Ok)
        self.buttonBox.setObjectName("buttonBox")
        self.gridLayout.addWidget(self.buttonBox, 2, 0, 1, 1)

        self.retranslateUi()
        self.buttonBox.accepted.connect(self.accept)  # Connect the accept signal
        self.buttonBox.rejected.connect(self.reject)  # Connect the reject signal
        self.add_user_btn.clicked.connect(self.add_user)  # Connect the add user button
        self.edit_user_btn.clicked.connect(self.edit_user)  # Connect the edit user button
        self.save_user_btn.clicked.connect(self.save_user)  # Connect the save user button
        self.admin_users.itemSelectionChanged.connect(self.load_selected_user_data)  # Load selected user data
        QtCore.QMetaObject.connectSlotsByName(self)

        # Load data into the table
        self.load_data()

    def retranslateUi(self):
        _translate = QtCore.QCoreApplication.translate
        self.setWindowTitle(_translate("AdminUsersDialog", "إدارة المستخدمين"))
        self.code.setText(_translate("AdminUsersDialog", "الكود :"))
        self.user_name.setText(_translate("AdminUsersDialog", "الاسم:"))
        self.add_user_btn.setText(_translate("AdminUsersDialog", "اضافة"))
        self.edit_user_btn.setText(_translate("AdminUsersDialog", "تعديل"))
        self.save_user_btn.setText(_translate("AdminUsersDialog", "حفظ"))

    def load_data(self):
        # Connect to the database and fetch data from the admin_users table
        conn = sqlite3.connect('agents.db')
        cursor = conn.cursor()
        cursor.execute("SELECT user_name, user_code FROM admin_users")
        rows = cursor.fetchall()
        conn.close()

        # Populate the QTableWidget with the data
        self.admin_users.setRowCount(len(rows))
        for row_index, row_data in enumerate(rows):
            self.admin_users.setItem(row_index, 0, QtWidgets.QTableWidgetItem(row_data[0]))
            self.admin_users.setItem(row_index, 1, QtWidgets.QTableWidgetItem(str(row_data[1])))

    def load_selected_user_data(self):
        # Load the selected user data into the input fields
        selected_row = self.admin_users.currentRow()
        if selected_row >= 0:
            user_name_item = self.admin_users.item(selected_row, 0)
            user_code_item = self.admin_users.item(selected_row, 1)

            if user_name_item and user_code_item:
                self.user_name_input.setText(user_name_item.text())
                self.code_input.setText(user_code_item.text())
                self.edit_user_btn.setEnabled(True)  # Enable the edit button
                self.save_user_btn.setEnabled(False)  # Disable the save button
                self.add_user_btn.setEnabled(False)  # Disable the add button

    def edit_user(self):
        # Enable editing
        self.user_name_input.setEnabled(True)
        self.code_input.setEnabled(True)
        self.save_user_btn.setEnabled(True)  # Enable the save button
        self.edit_user_btn.setEnabled(False)  # Disable the edit button
        self.add_user_btn.setEnabled(False)  # Disable the add button

    def save_user(self):
        # Get the user input
        user_name = self.user_name_input.text()
        user_code = self.code_input.text()

        if user_name and user_code:
            try:
                # Connect to the database
                conn = sqlite3.connect('agents.db')
                cursor = conn.cursor()

                # Check if a row is selected for updating
                selected_row = self.admin_users.currentRow()
                if selected_row >= 0:
                    # Update existing user
                    original_code = self.admin_users.item(selected_row, 1).text()
                    cursor.execute("UPDATE admin_users SET user_name = ?, user_code = ? WHERE user_code = ?", 
                                   (user_name, user_code, original_code))
                    conn.commit()
                    conn.close()

                    # Update the table
                    self.load_data()

                    # Clear input fields
                    self.user_name_input.clear()
                    self.code_input.clear()

                    # Reset buttons
                    self.user_name_input.setEnabled(False)
                    self.code_input.setEnabled(False)
                    self.save_user_btn.setEnabled(False)  # Disable the save button
                    self.edit_user_btn.setEnabled(True)  # Enable the edit button
                    self.add_user_btn.setEnabled(True)  # Enable the add button

                    # Show success message
                    QtWidgets.QMessageBox.information(self, "نجاح", "تم حفظ التغييرات بنجاح.")
            except sqlite3.Error as e:
                # Show error message
                QtWidgets.QMessageBox.critical(self, "خطأ", f"حدث خطأ أثناء حفظ التغييرات: {str(e)}")
        else:
            # Show warning message if fields are empty
            QtWidgets.QMessageBox.warning(self, "تحذير", "يرجى ملء جميع الحقول.")

    def add_user(self):
        # Get the user input
        user_name = self.user_name_input.text()
        user_code = self.code_input.text()

        if user_name and user_code:
            try:
                # Connect to the database and insert a new user
                conn = sqlite3.connect('agents.db')
                cursor = conn.cursor()
                cursor.execute("INSERT INTO admin_users (user_name, user_code) VALUES (?, ?)", (user_name, user_code))
                conn.commit()
                conn.close()

                # Update the table
                self.load_data()

                # Clear input fields
                self.user_name_input.clear()
                self.code_input.clear()

                # Reset buttons
                self.save_user_btn.setEnabled(False)  # Disable the save button
                self.edit_user_btn.setEnabled(False)  # Disable the edit button
                self.add_user_btn.setEnabled(True)  # Enable the add button

                # Show success message
                QtWidgets.QMessageBox.information(self, "نجاح", "تمت إضافة المستخدم بنجاح.")
            except sqlite3.Error as e:
                # Show error message
                QtWidgets.QMessageBox.critical(self, "خطأ", f"حدث خطأ أثناء إضافة المستخدم: {str(e)}")
        else:
            # Show warning message if fields are empty
            QtWidgets.QMessageBox.warning(self, "تحذير", "يرجى ملء جميع الحقول.")
class EditAgentsDialog(QtWidgets.QDialog):
    def __init__(self, parent=None):
        super(EditAgentsDialog, self).__init__(parent)
        self.setupUi(self)

    def setupUi(self, Dialog):
        Dialog.setObjectName("Dialog")
        Dialog.resize(900, 700)
        self.gridLayout = QtWidgets.QGridLayout(Dialog)
        self.gridLayout.setObjectName("gridLayout")

        self.agents_section = QtWidgets.QTableWidget(Dialog)
        self.agents_section.setObjectName("agents_section")
        self.agents_section.setColumnCount(9)
        self.agents_section.setHorizontalHeaderLabels(
            [
                "الاسم",
                "الرتبة العسكرية",
                "الرقم الخاص",
                "الرقم العام",
                "السجل المدني",
                "رقم الجوال",
                "المجموعة",
                "المحافظة",
                "الإجازات المتبقية",
            ]
        )
        self.agents_section.setEditTriggers(QtWidgets.QAbstractItemView.DoubleClicked)
        self.agents_section.setLayoutDirection(QtCore.Qt.RightToLeft)  # Right-to-left
        self.gridLayout.addWidget(self.agents_section, 0, 0, 1, 2)

        self.pushButton = QtWidgets.QPushButton(Dialog)
        self.pushButton.setObjectName("pushButton")
        self.gridLayout.addWidget(self.pushButton, 1, 0, 1, 1)

        self.pushButton_2 = QtWidgets.QPushButton(Dialog)
        self.pushButton_2.setObjectName("pushButton_2")
        self.gridLayout.addWidget(self.pushButton_2, 1, 1, 1, 1)

        self.buttonBox = QtWidgets.QDialogButtonBox(Dialog)
        self.buttonBox.setOrientation(QtCore.Qt.Horizontal)
        self.buttonBox.setStandardButtons(QtWidgets.QDialogButtonBox.Cancel|QtWidgets.QDialogButtonBox.Ok)
        self.buttonBox.setObjectName("buttonBox")
        self.gridLayout.addWidget(self.buttonBox, 2, 0, 1, 2)

        self.retranslateUi(Dialog)
        self.buttonBox.accepted.connect(Dialog.accept)  # type: ignore
        self.buttonBox.rejected.connect(Dialog.reject)  # type: ignore
        QtCore.QMetaObject.connectSlotsByName(Dialog)

        # Additional setup
        self.initUI()
        self.populate_table()

    def retranslateUi(self, Dialog):
        _translate = QtCore.QCoreApplication.translate
        Dialog.setWindowTitle(_translate("Dialog", "تعديل جدول المستخدمين"))
        self.pushButton.setText(_translate("Dialog", "حذف المحدد"))
        self.pushButton_2.setText(_translate("Dialog", "حفظ التغييرات"))

    def initUI(self):
        # Connect the cellChanged signal to a custom slot
        self.agents_section.cellChanged.connect(self.cell_changed)

        # Connect buttons to their respective functions
        self.pushButton.clicked.connect(self.delete_selected)
        self.pushButton_2.clicked.connect(self.save_changes)

    def get_users_data(self):
        connection = sqlite3.connect("agents.db")
        cursor = connection.cursor()

        cursor.execute(
            "SELECT name, military_rank, military_number, general_number, civil_registry, mobile_number, the_group, governorate, the_remaining_holidays FROM users"
        )
        data = cursor.fetchall()

        connection.close()
        return data

    def populate_table(self):
        data = self.get_users_data()
        self.agents_section.setRowCount(len(data))

        for row_num, row_data in enumerate(data):
            for col_num, col_data in enumerate(row_data):
                item = QtWidgets.QTableWidgetItem(str(col_data))
                if col_num == 2:  # Make the "Military Number" column non-editable
                    item.setFlags(item.flags() & ~QtCore.Qt.ItemIsEditable)
                self.agents_section.setItem(row_num, col_num, item)

    def cell_changed(self, row, column):
        item = self.agents_section.item(row, column)
        item.setBackground(QtGui.QColor("#FFEB3B"))  # Highlight the edited cell

    def save_changes(self):
        connection = sqlite3.connect("agents.db")
        cursor = connection.cursor()

        row_count = self.agents_section.rowCount()
        col_count = self.agents_section.columnCount()

        for row in range(row_count):
            row_data = []
            for column in range(col_count):
                item = self.agents_section.item(row, column)
                row_data.append(item.text() if item else "")

            cursor.execute(
                """
                UPDATE users
                SET name=?, military_rank=?, general_number=?, civil_registry=?, mobile_number=?, the_group=?, governorate=?, the_remaining_holidays=?
                WHERE military_number=?
            """,
                (
                    row_data[0],
                    row_data[1],
                    row_data[3],
                    row_data[4],
                    row_data[5],
                    row_data[6],
                    row_data[7],
                    row_data[8],
                    row_data[2],
                ),
            )

        connection.commit()
        connection.close()

        # Reset the background color of all cells
        for row in range(row_count):
            for column in range(col_count):
                item = self.agents_section.item(row, column)
                item.setBackground(QtGui.QColor("white"))
        

    def delete_selected(self):
        selected_items = self.agents_section.selectedItems()
        if not selected_items:
            QtWidgets.QMessageBox.warning(
                self, "لا يوجد تحديد", "يرجى تحديد صف واحد على الأقل للحذف."
            )
            return

        connection = sqlite3.connect("agents.db")
        cursor = connection.cursor()

        # Extract unique row indices from the selected items
        rows_to_delete = set(item.row() for item in selected_items)

        for row_number in sorted(rows_to_delete, reverse=True):
            military_number_item = self.agents_section.item(row_number, 2)
            if military_number_item:
                military_number = military_number_item.text()
                cursor.execute(
                    "DELETE FROM users WHERE military_number=?", (military_number,)
                )
                self.agents_section.removeRow(row_number)

        connection.commit()
        connection.close()
##########################################################################
class Ui_MainWindow(object):
    ############################ Security Functions ##########################
    def check_admin_user(self):
        input_dialog = QtWidgets.QInputDialog(self)
        input_dialog.setWindowTitle("التحقق من هوية المستخدم")
        input_dialog.setLabelText("ادخل الكود الخاص بك:")
        input_dialog.setTextEchoMode(QtWidgets.QLineEdit.Password)
        input_dialog.setTextEchoMode(QtWidgets.QLineEdit.Password)
        ok = input_dialog.exec_()
        user_code = input_dialog.textValue()

        print(user_code)
        if ok:
            user_name = self.get_admin_user(user_code)
            if user_name:
                QtWidgets.QMessageBox.information(
                    self,
                    "تم الاضافة بنجاح",
                    f"المستخدم : {user_name}",
                )
                return user_code
            else:
                QtWidgets.QMessageBox.warning(
                    self,
                    "تعذر الاضافة",
                    f"لم يتم العثور علي المتسخدم",
                )
                return False
        else:
            return False

    def check_admin(self):
        input_dialog = QtWidgets.QInputDialog(self)
        input_dialog.setWindowTitle("التحقق من هوية المستخدم")
        input_dialog.setLabelText("ادخل الكود الخاص بك:")
        input_dialog.setTextEchoMode(QtWidgets.QLineEdit.Password)
        input_dialog.setTextEchoMode(QtWidgets.QLineEdit.Password)
        ok = input_dialog.exec_()
        user_code = input_dialog.textValue()

        print(user_code)
        if ok:
            user_name = self.get_admin(user_code)
            if user_name:
                QtWidgets.QMessageBox.information(
                    self,
                    "تم الاضافة بنجاح",
                    f"شكرا لك {user_name}",
                )
                return user_code
            else:
                QtWidgets.QMessageBox.warning(
                    self,
                    " خطاء في الادخال",
                    f"انت لا تمتلك الصلاحية    ",
                )
                return False
        else:
            return False
    ##########################################################################
    ############################ Menu Bar Functions ##########################
    def HolidayReset(self):
        check_admin = self.check_admin()
        if check_admin != False:
            self.reset_remaining_holidays() 
            msg_box = QtWidgets.QMessageBox()
            msg_box.setIcon(QtWidgets.QMessageBox.Warning)
            msg_box.setWindowTitle(" تحديث الاجازات")
            msg_box.setText("تم تحديث الاجازات بنجاح")
            msg_box.setStandardButtons(QtWidgets.QMessageBox.Ok)
            msg_box.exec_()
    def report_bydate(self):
        check_admin = self.check_admin()
        if check_admin != False:
            self.edit_window = report_bydate()
            self.edit_window.exec_()
    def report(self):
        check_admin = self.check_admin()
        if check_admin != False:
            self.edit_window = report()
            self.edit_window.exec_()
    def UpdateAdmin(self):
        check_admin = self.check_admin()
        if check_admin != False:
            self.edit_window = UpdateAdmin()
            self.edit_window.exec_()
    def AdminUsersDialog(self):
        check_admin = self.check_admin()
        if check_admin != False:
            self.edit_window = AdminUsersDialog()
            self.edit_window.exec_()
    ##########################################################################
    ############################  Absent Functions ###########################
    def fetch_current_absentees(self):
        conn = sqlite3.connect("agents.db")
        cursor = conn.cursor()

        # Get the current Gregorian date and convert it to Hijri
        current_date_gregorian = datetime.now().date()
        current_date_hijri = convert.Gregorian(
            current_date_gregorian.year,
            current_date_gregorian.month,
            current_date_gregorian.day,
        ).to_hijri()
        current_date_hijri_str = f"{current_date_hijri.year}-{current_date_hijri.month:02d}-{current_date_hijri.day:02d}"

        query = """
            SELECT military_number, return_date, check_in_date
            FROM holidays_history
            WHERE Return_date < ?
            AND (check_in_date IS NULL OR check_in_date = '');
        """

        cursor.execute(query, (current_date_hijri_str,))
        absentees = cursor.fetchall()

        filtered_absentees = []

        for absentee in absentees:
            military_number, return_date_hijri, check_in_date_hijri = absentee

            if return_date_hijri:
                return_date_hijri_obj = convert.Hijri(
                    *map(int, return_date_hijri.split("-"))
                )
                return_date_gregorian = return_date_hijri_obj.to_gregorian()
                return_date_gregorian_date = date(
                    return_date_gregorian.year,
                    return_date_gregorian.month,
                    return_date_gregorian.day,
                )
            else:
                return_date_gregorian_date = None

            current_date_gregorian = datetime.now().date()

            if return_date_gregorian_date:
                duration_of_absence = (
                    current_date_gregorian - return_date_gregorian_date
                ).days
            else:
                duration_of_absence = None
            print(duration_of_absence)
            if duration_of_absence == 1:
                filtered_absentees.append(
                    {
                        "military_number": military_number,
                        "return_date": return_date_hijri,
                        "check_in_date": check_in_date_hijri,
                    }
                )

        conn.close()

        return filtered_absentees
    def alert_new_absentees(self):
        absentees = self.fetch_current_absentees()
        print(type(absentees))  # For debugging purposes
        print(absentees)
        if absentees:
            absent_message = "تحقق من قائمة المطوف\n"
            for absentee in absentees:
                print(absentee["military_number"])
                military_number=absentee["military_number"]
                absent_message += (
                    f"الاسم : {self.get_name_by_military_number(military_number)}\n"
                )

            # Display the alert in Arabic
            msg_box = QtWidgets.QMessageBox()
            msg_box.setIcon(QtWidgets.QMessageBox.Warning)
            msg_box.setWindowTitle("تنبيه المطوف")
            msg_box.setText(absent_message)
            msg_box.setStandardButtons(QtWidgets.QMessageBox.Ok)
            msg_box.exec_()
        else:
            # Optional: Display a message if no absentees are found, in Arabic
            msg_box = QtWidgets.QMessageBox()
            msg_box.setIcon(QtWidgets.QMessageBox.Information)
            msg_box.setWindowTitle("تنبيه المطوف")
            msg_box.setText("لم يتم اكتشاف أي تطويف جديد.")
            msg_box.setStandardButtons(QtWidgets.QMessageBox.Ok)
            msg_box.exec_()

    def add_data_to_table(self, military_number, duration):
        # Get the row position to insert the new data
        row_position = self.alert_section.rowCount()
        self.alert_section.insertRow(row_position)
        
        # Retrieve the name using the military number
        name = self.get_name_by_military_number(military_number)
        
        # Create the QTableWidgetItem for the military number, name, and duration
        military_number_item = QtWidgets.QTableWidgetItem(military_number)
        name_item = QtWidgets.QTableWidgetItem(name)
        duration_item = QtWidgets.QTableWidgetItem(duration)
        
        # Set the items in the table row
        self.alert_section.setItem(row_position, 0, military_number_item)
        self.alert_section.setItem(row_position, 1, name_item)
        self.alert_section.setItem(row_position, 2, duration_item)

    def check_absent(self):
        connection = sqlite3.connect("agents.db")
        cursor = connection.cursor()

        cursor.execute(
            "SELECT military_number, return_date, check_in_date FROM holidays_history"
        )
        data = cursor.fetchall()

        today_hijri = self.today_date()
        today_hijri_str = (
            f"{today_hijri.year}-{today_hijri.month:02d}-{today_hijri.day:02d}"
        )

        absentees = []

        for row in data:
            military_number, return_date, check_in_date = row
            if not check_in_date:  # Only process if check_in_date is empty
                if return_date:
                    days_diff = self.days_between_hijri_dates(
                        return_date, today_hijri_str
                    )
                    if days_diff > 0:
                        absentees.append((military_number, days_diff))
                    elif days_diff == 1:
                        absentees.append((military_number, days_diff))

        connection.close()
        return absentees
    def separate_absentees(self, data):
        military_numbers = []
        days_absent = []

        for military_number, days_absent_count in data:
            military_numbers.append(military_number)
            days_absent.append(days_absent_count)

        return military_numbers, days_absent

    ##########################################################################
    ############################ Runtime Functions ###########################
    def check_and_display_absents(self):
        absentees = self.check_absent()
        military_numbers, days_absent = self.separate_absentees(absentees)

        # Clear the table before adding new data
        self.alert_section.setRowCount(0)

        for military_number, days in zip(military_numbers, days_absent):
            self.add_data_to_table(str(military_number), str(days))


    def update_addholiday_section(self):
        self.Holiday_End_input.setText(
            str(self.get_future_date(self.Holiday_Duration_input.value()))
        )
    def populate_table(self):
        data = self.get_users_list()
        self.agents_section.setRowCount(len(data))

        for row_num, row_data in enumerate(data):
            for col_num, col_data in enumerate(row_data):
                self.agents_section.setItem(
                    row_num, col_num, QtWidgets.QTableWidgetItem(str(col_data))
                )
    ##########################################################################
    ############################  Side Functions  ############################ 
    def reset_remaining_holidays(self):
        conn = sqlite3.connect("agents.db")
        cursor = conn.cursor()

        # Update the_remaining_holidays to 60 for all users
        update_query = "UPDATE users SET the_remaining_holidays = 60"
        cursor.execute(update_query)

        # Commit the changes and close the connection
        conn.commit()
        conn.close()
    
    def get_admin_user(self, user_code):
        # Connect to the database
        conn = sqlite3.connect("agents.db")
        cursor = conn.cursor()

        # SQL statement to check for the user_code
        query = """
        SELECT user_name FROM admin_users WHERE user_code = ?;
        """

        # Execute the query
        cursor.execute(query, (user_code,))
        result = cursor.fetchone()

        # Close the database connection
        conn.close()

        if result:
            return result[0]  # Return the user_name
        else:
            return None

    def get_admin(self, user_code):
        # Connect to the database
        conn = sqlite3.connect("agents.db")
        cursor = conn.cursor()

        # SQL statement to check for the user_code
        query = """
            SELECT username FROM admin WHERE password = ?;
            """

        # Execute the query
        cursor.execute(query, (user_code,))
        result = cursor.fetchone()

        # Close the database connection
        conn.close()

        if result:
            return result[0]  # Return the user_name
        else:
            return None
    def edit_remaining_holidays(self, military_number, days):
        # Connect to the database
        conn = sqlite3.connect("agents.db")
        cursor = conn.cursor()

        try:
            # SQL statement to update the_remaining_holidays column
            update_sql = """
            UPDATE users
            SET the_remaining_holidays = the_remaining_holidays - ?
            WHERE military_number = ?
            """

            cursor.execute(update_sql, (days, military_number))

            conn.commit()

            print("Remaining holidays updated successfully!")

        except sqlite3.Error as e:
            print(f"Error: {e}")

        finally:
            cursor.close()
            conn.close()

    def get_name_by_military_number(self, military_number):
        conn = sqlite3.connect("agents.db")
        cursor = conn.cursor()
        query = "SELECT name FROM users WHERE military_number = ?"
        cursor.execute(query, (military_number,))
        result = cursor.fetchone()
        conn.close()

        if result:
            return result[0]
        else:
            return "Name not found"

    def update_check_in_date(self, military_number, start_date, return_date):
        # Convert start_date and return_date to string format
        start_date_str = start_date
        return_date_str = return_date

        # Convert current date to Hijri
        today = QtCore.QDate.currentDate()
        hijri_date = convert.Gregorian(today.year(), today.month(), today.day()).to_hijri()
        hijri_date_str = f"{hijri_date.year:04}-{hijri_date.month:02}-{hijri_date.day:02}"

        # Connect to the database
        conn = sqlite3.connect("agents.db")
        cursor = conn.cursor()

        # SQL statement to search for the military number, start date, and return date
        search_query = """
        SELECT * FROM holidays_history
        WHERE military_number = ?
        AND start_Date = ?
        AND return_Date = ?;
        """
        cursor.execute(search_query, (military_number, start_date_str, return_date_str))
        results = cursor.fetchall()

        if results:
            # Record found; update the check_in_date with the current Hijri date
            update_query = """
            UPDATE holidays_history
            SET check_in_date = ?
            WHERE military_number = ?
            AND start_Date = ?
            AND return_Date = ?;
            """
            cursor.execute(update_query, (hijri_date_str, military_number, start_date_str, return_date_str))
            conn.commit()

        # Close the database connection
        conn.close()
    def calculate_all_durations_of_absence(self):
        # Connect to the database
        conn = sqlite3.connect("agents.db")
        cursor = conn.cursor()
        # SQL statement to get all military numbers, return_date, and check_in_date
        query = """
        SELECT military_number, return_date, check_in_date 
        FROM holidays_history;
        """

        # Execute the query
        cursor.execute(query)
        results = cursor.fetchall()

        for result in results:
            military_number, return_date_hijri, check_in_date_hijri = result

            # Convert Hijri dates to Gregorian
            if return_date_hijri:
                return_date_hijri_obj = convert.Hijri(
                    *map(int, return_date_hijri.split("-"))
                )
                return_date_gregorian = return_date_hijri_obj.to_gregorian()
                return_date_gregorian_date = date(
                    return_date_gregorian.year,
                    return_date_gregorian.month,
                    return_date_gregorian.day,
                )
            else:
                return_date_gregorian_date = None

            # If check_in_date is empty, set it to today's date
            if check_in_date_hijri:
                check_in_date_hijri_obj = convert.Hijri(
                    *map(int, check_in_date_hijri.split("-"))
                )
                check_in_date_gregorian = check_in_date_hijri_obj.to_gregorian()
                check_in_date_gregorian_date = date(
                    check_in_date_gregorian.year,
                    check_in_date_gregorian.month,
                    check_in_date_gregorian.day,
                )
            else:
                current_date_gregorian = datetime.now().date()
                check_in_date_hijri_obj = convert.Gregorian(
                    current_date_gregorian.year,
                    current_date_gregorian.month,
                    current_date_gregorian.day,
                ).to_hijri()
                check_in_date_hijri = f"{check_in_date_hijri_obj.year}-{check_in_date_hijri_obj.month:02d}-{check_in_date_hijri_obj.day:02d}"
                check_in_date_gregorian_date = current_date_gregorian

            # Calculate the duration of absence
            if return_date_gregorian_date and check_in_date_gregorian_date:
                duration_of_absence = (
                    check_in_date_gregorian_date - return_date_gregorian_date
                ).days
            else:
                duration_of_absence = None

            # Update the duration_of_absence and possibly check_in_date in the database
            update_query = """
            UPDATE holidays_history
            SET duration_of_absence = ?
            WHERE military_number = ? AND return_date = ?;
            """
            cursor.execute(
                update_query,
                (
                    duration_of_absence,
                    military_number,
                    return_date_hijri,
                ),
            )

        # Commit the changes and close the database connection
        conn.commit()
        conn.close()

        print("Duration of absence updated for all records.")
    def check_absentees(self, military_number):
        # Connect to the database
        conn = sqlite3.connect("agents.db")
        cursor = conn.cursor()

        # SQL statement to retrieve relevant data using military_number
        query = """
        SELECT the_kind_of_holiday, duration_of_vacation, Start_Date, Return_date, duration_of_absence, user_code, check_in_date 
        FROM holidays_history
        WHERE military_number = ?;
        """

        # Execute the query and fetch all results
        cursor.execute(query, (military_number,))
        results = cursor.fetchall()

        # Close the database connection
        conn.close()

        holidays = []

        # Process all holidays for the user
        for result in results:
            (
                the_kind_of_holiday,
                duration_of_vacation,
                start_date_hijri,
                return_date_hijri,
                duration_of_absence,
                user_code,
                check_in_date,
            ) = result

            # If check_in_date is None, calculate the remaining days
            if check_in_date is None:
                # Convert Hijri dates to Gregorian
                if return_date_hijri:
                    return_date_hijri_obj = convert.Hijri(*map(int, return_date_hijri.split("-")))
                    return_date_gregorian = return_date_hijri_obj.to_gregorian()
                    return_date_gregorian_date = date(
                        return_date_gregorian.year,
                        return_date_gregorian.month,
                        return_date_gregorian.day,
                    )
                else:
                    return_date_gregorian_date = None

                # Calculate the remaining days
                current_date_gregorian = datetime.now().date()
                if return_date_gregorian_date:
                    remaining_days = (return_date_gregorian_date - current_date_gregorian).days
                    if remaining_days < 0:
                        remaining_days = 0
                else:
                    remaining_days = None

                holidays.append({
                    "current_date_hijri": convert.Gregorian(
                        current_date_gregorian.year,
                        current_date_gregorian.month,
                        current_date_gregorian.day,
                    ).to_hijri(),
                    "remaining_days": remaining_days,
                    "the_kind_of_holiday": the_kind_of_holiday,
                    "duration_of_vacation": duration_of_vacation,
                    "start_Date": start_date_hijri,
                    "return_date": return_date_hijri,
                    "duration_of_absence": duration_of_absence,
                    "user_code": user_code,
                    "check_in_date": check_in_date if check_in_date else "Not checked in",
                })

        if holidays:
            print("------------------------------------------------")
            print(holidays)
            print("------------------------------------------------")
            return holidays
        else:
            return "theres no holdays"
            return False

    def current_holiday(self, military_number):
        while_holiday = self.check_current_holiday(military_number)

        unchecked_holiday = self.check_absentees(military_number)
        print("unchecked_holiday="+str(unchecked_holiday))
        if unchecked_holiday != False:
            try:
                    currnet_holiday = unchecked_holiday[0]
                    self.Current_Holiday_Kind.setText(
                                f"نوع الاجازة : {currnet_holiday['the_kind_of_holiday']}"
                            )
                    self.Current_Holiday_Duration.setText(
                                f"مدة الاجازة : {currnet_holiday['duration_of_vacation']}"
                            )
                    self.Current_Holiday_Start.setText(
                                f"تاريخ البدء : {currnet_holiday['start_Date']}"
                            )
                    self.Current_Holiday_End.setText(
                                f"تاريخ الانتهاء : {currnet_holiday['return_date']}"
                            )
                    self.Absence_Period.setText(
                                f"فترة المطوف : {currnet_holiday['duration_of_absence']}"
                            )
                    self.Remaining_Days.setText(
                                f"الايام المتبقية : {currnet_holiday['remaining_days']}"
                            )
            except TypeError:
                    print(TypeError)
                    self.Current_Holiday_Kind.setText(f"نوع الاجازة : ")
                    self.Current_Holiday_Duration.setText(f"مدة الاجازة :")
                    self.Current_Holiday_Start.setText(f"تاريخ البدء : ")
                    self.Current_Holiday_End.setText(f"تاريخ الانتهاء :")
                    self.Absence_Period.setText(f"فترة المطوف : ")
                    self.Remaining_Days.setText(f"الايام المتبقية : ")
        else:
            if while_holiday == False:
                    currnet_holiday = while_holiday
                    self.Current_Holiday_Kind.setText(
                                f"نوع الاجازة : {currnet_holiday['the_kind_of_holiday']}"
                            )
                    self.Current_Holiday_Duration.setText(
                                f"مدة الاجازة : {currnet_holiday['duration_of_vacation']}"
                            )
                    self.Current_Holiday_Start.setText(
                                f"تاريخ البدء : {currnet_holiday['start_Date']}"
                            )
                    self.Current_Holiday_End.setText(
                                f"تاريخ الانتهاء : {currnet_holiday['return_date']}"
                            )
                    self.Absence_Period.setText(
                                f"فترة المطوف : {currnet_holiday['duration_of_absence']}"
                            )
                    self.Remaining_Days.setText(
                                f"الايام المتبقية : {currnet_holiday['remaining_days']}"
                            )
            else:
                self.Current_Holiday_Kind.setText(f"نوع الاجازة : ")
                self.Current_Holiday_Duration.setText(f"مدة الاجازة : ")
                self.Current_Holiday_Start.setText(f"تاريخ البدء : ")
                self.Current_Holiday_End.setText(f"تاريخ الانتهاء : ")
                self.Absence_Period.setText(f"فترة المطوف : ")
                self.Remaining_Days.setText(f"الايام المتبقية : ")

    def check_current_holiday(self, military_number):
        # Connect to the database
        conn = sqlite3.connect("agents.db")
        cursor = conn.cursor()

        # Get the current Gregorian date and convert it to Hijri
        current_date_gregorian = datetime.now().date()
        current_date_hijri = convert.Gregorian(
            current_date_gregorian.year,
            current_date_gregorian.month,
            current_date_gregorian.day,
        ).to_hijri()
        current_date_hijri_str = f"{current_date_hijri.year}-{current_date_hijri.month:02d}-{current_date_hijri.day:02d}"

        # SQL statement to check for a running holiday or a holiday without a check_in_date
        query = """
            SELECT the_kind_of_holiday, duration_of_vacation, Start_Date, Return_date, duration_of_absence, user_code, check_in_date
            FROM holidays_history
            WHERE military_number = ?
            AND Start_Date <= ?
            AND Return_date >= ?
            AND (check_in_date IS NULL OR check_in_date = '');
            """

        # Execute the query
        cursor.execute(
            query, (military_number, current_date_hijri_str, current_date_hijri_str)
        )
        result = cursor.fetchone()

        # Close the database connection
        conn.close()

        if result:
            (
                the_kind_of_holiday,
                duration_of_vacation,
                start_date_hijri,
                return_date_hijri,
                duration_of_absence,
                user_code,
                check_in_date,
            ) = result

            return_date_hijri_obj = convert.Hijri(
                *map(int, return_date_hijri.split("-"))
            )
            return_date_gregorian = return_date_hijri_obj.to_gregorian()
            return_date_gregorian_date = date(
                return_date_gregorian.year,
                return_date_gregorian.month,
                return_date_gregorian.day,
            )
            remaining_days = (return_date_gregorian_date - current_date_gregorian).days
            if remaining_days < 0:
                remaining_days = 0
                return remaining_days
            # Set the checkbox state based on check_in_date
            
            return {
                "current_date_hijri": current_date_hijri_str,
                "remaining_days": remaining_days,
                "the_kind_of_holiday": the_kind_of_holiday,
                "duration_of_vacation": duration_of_vacation,
                "start_Date": start_date_hijri,
                "return_date": return_date_hijri,
                "duration_of_absence": duration_of_absence,
                "user_code": user_code,
                "check_in_date": check_in_date,
            }
        else:
            return False
    def get_username(self, user_code):
        conn = sqlite3.connect("agents.db")
        cursor = conn.cursor()
        cursor.execute(
            "SELECT user_name FROM admin_users WHERE user_code = ?", (user_code,)
        )
        result = cursor.fetchone()
        conn.close()
        return result[0] if result else "غير معروف"
    def search_holidays_by_military_number(self, military_number):
        # Convert current date to Hijri
        today = QtCore.QDate.currentDate()

        hijri_year = (
            convert.Gregorian(today.year(), today.month(), today.day()).to_hijri().year
        )

        hijri_year = str(hijri_year)

        # Connect to the database
        conn = sqlite3.connect("agents.db")
        cursor = conn.cursor()

        # SQL statement to search for the military number and filter by Hijri year
        search_query = """
        SELECT * FROM holidays_history
        WHERE military_number = ?
        AND strftime('%Y', start_Date) = ?;
        """
        cursor.execute(search_query, (military_number, hijri_year))
        results = cursor.fetchall()

        # Replace user_code with username
        holidays_with_names = []
        for row in results:
            user_code = row[6]  # Assuming user_code is at index 6
            username = self.get_username(user_code)
            holidays_with_names.append(
                row[:6] + (username,) + row[7:]  # Replace user_code with username
            )

        # Close the database connection
        conn.close()

        return holidays_with_names
    def holiday_history(self):
        try:
            results = self.search_holidays_by_military_number(self.Military_Code.text())
            self.Holidays_History.setRowCount(
                len(results)
            )  # Set the row count to the number of results
            for row_num, row_data in enumerate(results):
                for col_num, data in enumerate(row_data):
                    self.Holidays_History.setItem(
                        row_num, col_num, QtWidgets.QTableWidgetItem(str(data))
                    )
        except Exception as e:
            pass
    def search_user(self,military_number=None, name=None):
        conn = sqlite3.connect('agents.db')
        cursor = conn.cursor()

        result = None  # Initialize result to None

        # Search by military_number first
        if military_number:
            cursor.execute("""
                SELECT name, military_rank, military_number, general_number, civil_registry, 
                    mobile_number, the_group, governorate, the_remaining_holidays, user_code
                FROM users
                WHERE military_number = ?
            """, (military_number,))
            result = cursor.fetchone()

        # If no result, search by name
        if not result and name:
            cursor.execute("""
                SELECT name, military_rank, military_number, general_number, civil_registry, 
                    mobile_number, the_group, governorate, the_remaining_holidays, user_code
                FROM users
                WHERE name = ?
            """, (name,))
            result = cursor.fetchone()

        conn.close()

        # Return data as a dictionary if found
        if result:
            keys = ["name", "military_rank", "military_number", "general_number", "civil_registry",
                    "mobile_number", "the_group", "governorate", "the_remaining_holidays", "user_code"]
            return dict(zip(keys, result))
        else:
            return None

        conn = sqlite3.connect('agents.db')
        cursor = conn.cursor()

        # Search by military_number first
        if military_number:
            cursor.execute("""
                SELECT name, military_rank, military_number, general_number, civil_registry, 
                    mobile_number, the_group, governorate, the_remaining_holidays, user_code
                FROM users
                WHERE military_number = ?
            """, (military_number,))
            result = cursor.fetchone()

        # If no result, search by name
        if not result and name:
            cursor.execute("""
                SELECT name, military_rank, military_number, general_number, civil_registry, 
                    mobile_number, the_group, governorate, the_remaining_holidays, user_code
                FROM users
                WHERE name = ?
            """, (name,))
            result = cursor.fetchone()

        conn.close()

        # Return data as a dictionary if found
        if result:
            keys = ["name", "military_rank", "military_number", "general_number", "civil_registry",
                    "mobile_number", "the_group", "governorate", "the_remaining_holidays", "user_code"]
            return dict(zip(keys, result))
        else:
            return None
    def get_users_list(self):
        # Connect to the SQLite database
        connection = sqlite3.connect("agents.db")  # Change to your database file name
        cursor = connection.cursor()

        # Fetch all data from the users table
        cursor.execute("SELECT * FROM users")
        data = cursor.fetchall()

        # Close the connection
        connection.close()
        return data
    ##########################################################################
    ############################  Date Functions  ############################
    
    def days_between_hijri_dates(self, start_date_str, end_date_str):
        start_date = convert.Hijri(*map(int, start_date_str.split("-"))).to_gregorian()
        end_date = convert.Hijri(*map(int, end_date_str.split("-"))).to_gregorian()
        delta = end_date - start_date
        return delta.days


    def today_date(self):
        # Get the current Gregorian date
        current_date = datetime.now()
        hijri_date = convert.Gregorian(
            current_date.year, current_date.month, current_date.day
        ).to_hijri()
        return hijri_date
    def get_future_date(self, days):
        # Get the current Gregorian date
        current_date = datetime.now()
        hijri_date_str = self.Holiday_Start_input.text()
        try:
            year, month, day = map(int, hijri_date_str.split("-"))
            gregorian_date = convert.Hijri(year, month, day).to_gregorian()
        except ValueError:
            gregorian_date = current_date
            return gregorian_date
        current_date = gregorian_date
        # Convert the current Gregorian date to Hijri date
        hijri_date = convert.Gregorian(
            current_date.year, current_date.month, current_date.day
        ).to_hijri()

        # Calculate the future Gregorian date
        future_date = current_date + timedelta(days=days)

        # Convert the future Gregorian date to Hijri date
        future_hijri_date = convert.Gregorian(
            future_date.year, future_date.month, future_date.day
        ).to_hijri()

        return future_hijri_date
    ##########################################################################
    ############################  Main Functions  ############################
    def PrintDialog(self,military_code, name):
        self.edit_window = PrintDialog("agents.db", military_code, name)
        self.edit_window.exec_()
    def EditHolidays(self,military_code):
        check_admin = self.check_admin()
        if check_admin != False:
            self.edit_window = EditHistory(self.Military_Code.text())
            self.edit_window.exec_()
    def EditAgent(self):
        check_admin = self.check_admin()
        if check_admin != False:
            self.edit_window = EditAgentsDialog()
            self.edit_window.exec_()
            
    def comeback(self):
        military_number = self.Military_Code.text()
        while_holiday = self.check_current_holiday(military_number)
        unchecked_holiday = self.check_absentees(military_number)
        print(unchecked_holiday)
        print(while_holiday)
        if unchecked_holiday != False:
            print(unchecked_holiday)
            print(while_holiday)
            if while_holiday == False:
                currnet_holiday = unchecked_holiday
                self.update_check_in_date(military_number, currnet_holiday['start_Date'], currnet_holiday['return_date'])

            else:
                currnet_holiday = while_holiday
                self.update_check_in_date(military_number, currnet_holiday['start_Date'], currnet_holiday['return_date'])
        else:
            pass
    def add_holiday(
    self,
    military_number,
    the_kind_of_holiday,
    duration_of_vacation,
    start_date,
    return_date,
):
        check_admin = self.check_admin_user()
        if check_admin != False:
            if military_number == "":
                msg = QMessageBox()
                msg.setIcon(QMessageBox.Information)
                msg.setWindowTitle("مشكلة")
                msg.setText("الرجاء ادخال كود المجند")
                msg.setStandardButtons(QMessageBox.Ok)
                msg.exec_()
            else:
                try:
                    conn = sqlite3.connect("agents.db")
                    cursor = conn.cursor()
                    duration_of_absence = "0"

                    try:
                        insert_sql = """
                        INSERT INTO holidays_history (military_number, the_kind_of_holiday, duration_of_vacation, start_date, duration_of_absence, return_date, user_code)
                        VALUES (?, ?, ?, ?, ?, ?, ?)
                        """
                        user_code = check_admin
                        cursor.execute(
                            insert_sql,
                            (
                                military_number,
                                the_kind_of_holiday,
                                duration_of_vacation,
                                start_date,
                                duration_of_absence,
                                return_date,
                                user_code,
                            ),
                        )
                        conn.commit()

                        # Show a success message box
                        msg = QMessageBox()
                        msg.setIcon(QMessageBox.Information)
                        msg.setWindowTitle("نجاح")
                        msg.setText("تم إضافة الإجازة بنجاح")
                        msg.setStandardButtons(QMessageBox.Ok)
                        msg.exec_()

                    except sqlite3.Error as e:
                        print(f"Error: {e}")

                    finally:
                        self.edit_remaining_holidays(military_number, duration_of_vacation)
                        cursor.close()
                        conn.close()

                except Exception as e:
                    print(f"Error: {e}")
                    return
    def Add_agent(
        self,
        military_number,
        name,
        general_number,
        military_rank,
        governorate,
        mobile_number,
        civil_registry,
        the_group,
    ):
        check_admin = self.check_admin()
        if check_admin != False:
            try:
                the_remaining_holidays = 60
                # Connect to the database
                conn = sqlite3.connect("agents.db")
                cursor = conn.cursor()
                
                if not military_number.isnumeric():
                    # Show an error message box if the military number is not numeric
                    msg = QMessageBox()
                    msg.setIcon(QMessageBox.Critical)
                    msg.setWindowTitle("خطأ")
                    msg.setText("يجب أن يكون الرقم الخاص رقميًا.")
                    msg.setStandardButtons(QMessageBox.Ok)
                    msg.exec_()
                    return

                # SQL statement to create the table if it doesn't exist
                create_table_sql = """
                CREATE TABLE IF NOT EXISTS users (
                    "military_number" NUMERIC,
                    "name" TEXT,
                    "general_number" NUMERIC,
                    "military_rank" TEXT NOT NULL,
                    "governorate" TEXT NOT NULL,
                    "mobile_number" NUMERIC NOT NULL,
                    "civil_registry" NUMERIC NOT NULL,
                    "the_group" TEXT NOT NULL,
                    "the_remaining_holidays" NUMERIC NOT NULL,
                    "user_code" NUMERIC,
                    PRIMARY KEY("military_number")
                );
                """
                # Execute the SQL statement to create the table
                cursor.execute(create_table_sql)

                # SQL statement to insert data into the table
                insert_data_sql = """
                INSERT INTO users (military_number, name, general_number, military_rank, governorate, mobile_number, civil_registry, the_group, the_remaining_holidays, user_code)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?);
                """

                user_code = check_admin
                # Execute the SQL statement to insert data
                cursor.execute(
                    insert_data_sql,
                    (
                        military_number,
                        name,
                        general_number,
                        military_rank,
                        governorate,
                        mobile_number,
                        civil_registry,
                        the_group,
                        the_remaining_holidays,
                        user_code,
                    ),
                )

                # Commit the changes and close the connection
                conn.commit()
                conn.close()

                # Show a success message box
                msg = QMessageBox()
                msg.setIcon(QMessageBox.Information)
                msg.setWindowTitle("نجاح")
                msg.setText("تم الاضافة بنجاح")
                msg.setStandardButtons(QMessageBox.Ok)
                msg.exec_()

            except Exception as e:
                print(f"Error: {e}")
                return
    def Search_Btn(self):
        try:
            self.calculate_all_durations_of_absence()
        except Exception as e:
            print(f"Error: {e}")
            return None
        military_number = self.Military_Code.text()
        name = self.lineEdit.text()
        agent_data = self.search_user(military_number, name)
        if agent_data:
            self.lineEdit.setText(agent_data['name'])
            self.Military_Code.setText(str(agent_data['military_number']))
            self.Agent_Data.setHtml(
                '<!DOCTYPE HTML PUBLIC "-//W3C//DTD HTML 4.0//EN" "http://www.w3.org/TR/REC-html40/strict.dtd">\n'
                '<html><head><meta name="qrichtext" content="1" /><style type="text/css">\n'
                "p, li { white-space: pre-wrap; }\n"
                "</style></head><body style=\" font-family:'MS Shell Dlg 2'; font-size:8.25pt; font-weight:400; font-style:normal;\">\n"
                f"<p align=\"right\" style=\" margin-top:0px; margin-bottom:0px; margin-left:0px; margin-right:0px; -qt-block-indent:0; text-indent:0px;\"><span style=\" font-size:10pt; font-weight:600;\">الاسم:{agent_data['name']}"
                f"<p align=\"right\" style=\" margin-top:0px; margin-bottom:0px; margin-left:0px; margin-right:0px; -qt-block-indent:0; text-indent:0px;\"><span style=\" font-size:10pt; font-weight:600;\">الرتبة : {agent_data['military_rank']}</span></p>"
                f"<p align=\"right\" style=\" margin-top:0px; margin-bottom:0px; margin-left:0px; margin-right:0px; -qt-block-indent:0; text-indent:0px;\"><span style=\" font-size:10pt; font-weight:600;\"> الرقم الخاص : {agent_data['military_number']}</span></p>\n"
                f"<p align=\"right\" style=\" margin-top:0px; margin-bottom:0px; margin-left:0px; margin-right:0px; -qt-block-indent:0; text-indent:0px;\"><span style=\" font-size:10pt; font-weight:600;\">الرقم العام : {agent_data['general_number']}</span></p>\n"
                f"<p align=\"right\" style=\" margin-top:0px; margin-bottom:0px; margin-left:0px; margin-right:0px; -qt-block-indent:0; text-indent:0px;\"><span style=\" font-size:10pt; font-weight:600;\">السجل : {agent_data['civil_registry']}</span></p>\n"
                f"<p align=\"right\" style=\" margin-top:0px; margin-bottom:0px; margin-left:0px; margin-right:0px; -qt-block-indent:0; text-indent:0px;\"><span style=\" font-size:10pt; font-weight:600;\">الجوال : {agent_data['mobile_number']} </span></p>\n"
                f"<p align=\"right\" style=\" margin-top:0px; margin-bottom:0px; margin-left:0px; margin-right:0px; -qt-block-indent:0; text-indent:0px;\"><span style=\" font-size:10pt; font-weight:600;\">المجموعة : {agent_data['the_group']}</span></p></body></html>"
                f"<p align=\"right\" style=\" margin-top:0px; margin-bottom:0px; margin-left:0px; margin-right:0px; -qt-block-indent:0; text-indent:0px;\"><span style=\" font-size:10pt; font-weight:600;\">الملاك : {agent_data['governorate']}</span></p>\n"
                f"<p align=\"right\" style=\" margin-top:0px; margin-bottom:0px; margin-left:0px; margin-right:0px; -qt-block-indent:0; text-indent:0px;\"><span style=\" font-size:10pt; font-weight:600;\"> المتبقي من الرصيد : {agent_data['the_remaining_holidays']} </span></p></body></html>"
                f"<p align=\"right\" style=\" margin-top:0px; margin-bottom:0px; margin-left:0px; margin-right:0px; -qt-block-indent:0; text-indent:0px;\"><span style=\" font-size:10pt; font-weight:600;\">اسم المستخدم : {self.get_username(agent_data['user_code'])}</span></p></body></html>"
            )
            self.current_holiday(military_number)
            self.holiday_history()
            
        else:
            QMessageBox.warning(self, "Error", "User not found.")
    ##########################################################################
    def setupUi(self, MainWindow):
        self.alert_new_absentees()
        MainWindow.setObjectName("MainWindow")
        MainWindow.resize(1083, 609)
        self.timer = QTimer(self)
        self.timer.timeout.connect(self.update_addholiday_section)
        self.timer.timeout.connect(self.populate_table)
        self.timer.timeout.connect(self.check_and_display_absents)
        self.timer.start(2000)
        self.centralwidget = QtWidgets.QWidget(MainWindow)
        self.centralwidget.setObjectName("centralwidget")
        self.gridLayout_5 = QtWidgets.QGridLayout(self.centralwidget)
        self.gridLayout_5.setObjectName("gridLayout_5")
        self.gridLayout = QtWidgets.QGridLayout()
        self.gridLayout.setObjectName("gridLayout")
        self.tabWidget = QtWidgets.QTabWidget(self.centralwidget)
        self.tabWidget.setLayoutDirection(QtCore.Qt.RightToLeft)
        self.tabWidget.setObjectName("tabWidget")
        self.main = QtWidgets.QWidget()
        self.main.setObjectName("main")
        self.gridLayout_6 = QtWidgets.QGridLayout(self.main)
        self.gridLayout_6.setObjectName("gridLayout_6")
        self.formLayout_2 = QtWidgets.QFormLayout()
        self.formLayout_2.setObjectName("formLayout_2")
        self.Currnet_Holiday_label = QtWidgets.QLabel(self.main)
        self.Currnet_Holiday_label.setObjectName("Currnet_Holiday_label")
        self.formLayout_2.setWidget(0, QtWidgets.QFormLayout.LabelRole, self.Currnet_Holiday_label)
        self.Current_Holiday_Kind = QtWidgets.QLabel(self.main)
        self.Current_Holiday_Kind.setObjectName("Current_Holiday_Kind")
        self.formLayout_2.setWidget(1, QtWidgets.QFormLayout.LabelRole, self.Current_Holiday_Kind)
        self.Current_Holiday_Duration = QtWidgets.QLabel(self.main)
        self.Current_Holiday_Duration.setObjectName("Current_Holiday_Duration")
        self.formLayout_2.setWidget(2, QtWidgets.QFormLayout.LabelRole, self.Current_Holiday_Duration)
        self.Current_Holiday_Start = QtWidgets.QLabel(self.main)
        self.Current_Holiday_Start.setObjectName("Current_Holiday_Start")
        self.formLayout_2.setWidget(3, QtWidgets.QFormLayout.LabelRole, self.Current_Holiday_Start)
        self.Current_Holiday_End = QtWidgets.QLabel(self.main)
        self.Current_Holiday_End.setObjectName("Current_Holiday_End")
        self.formLayout_2.setWidget(4, QtWidgets.QFormLayout.LabelRole, self.Current_Holiday_End)
        self.Absence_Period = QtWidgets.QLabel(self.main)
        self.Absence_Period.setObjectName("Absence_Period")
        self.formLayout_2.setWidget(5, QtWidgets.QFormLayout.LabelRole, self.Absence_Period)
        self.Remaining_Days = QtWidgets.QLabel(self.main)
        self.Remaining_Days.setObjectName("Remaining_Days")
        self.formLayout_2.setWidget(6, QtWidgets.QFormLayout.LabelRole, self.Remaining_Days)
        self.pushButton = QtWidgets.QPushButton(self.main)
        self.pushButton.setObjectName("pushButton")
        self.pushButton.clicked.connect(self.comeback)
        self.formLayout_2.setWidget(7, QtWidgets.QFormLayout.FieldRole, self.pushButton)
        self.gridLayout_6.addLayout(self.formLayout_2, 2, 1, 1, 1)
        self.gridLayout_2 = QtWidgets.QGridLayout()
        self.gridLayout_2.setObjectName("gridLayout_2")
        self.gridLayout_6.addLayout(self.gridLayout_2, 5, 1, 1, 2)
        self.gridLayout_3 = QtWidgets.QGridLayout()
        self.gridLayout_3.setObjectName("gridLayout_3")
        self.edit_1 = QtWidgets.QPushButton(self.main)
        font = QtGui.QFont()
        font.setPointSize(12)
        self.edit_1.setFont(font)
        self.edit_1.setObjectName("edit_1")
        self.edit_1.clicked.connect(self.EditHolidays)
        self.gridLayout_3.addWidget(self.edit_1, 2, 0, 1, 1)
        self.print_1 = QtWidgets.QPushButton(self.main)
        font = QtGui.QFont()
        font.setPointSize(12)
        self.print_1.setFont(font)
        self.print_1.setObjectName("print_1")
        self.gridLayout_3.addWidget(self.print_1, 3, 0, 1, 1)
        self.print_1.clicked.connect(lambda: self.PrintDialog(self.Military_Code.text(),self.lineEdit.text()))
        self.Holidays_History = QtWidgets.QTableWidget(self.main)
        self.Holidays_History.setRowCount(60)
        self.Holidays_History.setColumnCount(7)
        self.Holidays_History.setObjectName("Holidays_History")
        item = QtWidgets.QTableWidgetItem()
        self.Holidays_History.setHorizontalHeaderItem(0, item)
        item = QtWidgets.QTableWidgetItem()
        self.Holidays_History.setHorizontalHeaderItem(1, item)
        item = QtWidgets.QTableWidgetItem()
        self.Holidays_History.setHorizontalHeaderItem(2, item)
        item = QtWidgets.QTableWidgetItem()
        self.Holidays_History.setHorizontalHeaderItem(3, item)
        item = QtWidgets.QTableWidgetItem()
        self.Holidays_History.setHorizontalHeaderItem(4, item)
        item = QtWidgets.QTableWidgetItem()
        self.Holidays_History.setHorizontalHeaderItem(5, item)
        item = QtWidgets.QTableWidgetItem()
        self.Holidays_History.setHorizontalHeaderItem(6, item)
        item = QtWidgets.QTableWidgetItem()
        self.Holidays_History.setItem(0, 0, item)
        item = QtWidgets.QTableWidgetItem()
        self.Holidays_History.setItem(0, 1, item)
        item = QtWidgets.QTableWidgetItem()
        self.Holidays_History.setItem(0, 2, item)
        item = QtWidgets.QTableWidgetItem()
        self.Holidays_History.setItem(0, 3, item)
        self.gridLayout_3.addWidget(self.Holidays_History, 0, 0, 2, 1)
        self.gridLayout_6.addLayout(self.gridLayout_3, 2, 2, 3, 1)
        self.formLayout = QtWidgets.QFormLayout()
        self.formLayout.setObjectName("formLayout")
        self.Holiday_Duration = QtWidgets.QLabel(self.main)
        self.Holiday_Duration.setObjectName("Holiday_Duration")
        self.formLayout.setWidget(0, QtWidgets.QFormLayout.LabelRole, self.Holiday_Duration)
        self.Holiday_Duration_input = QtWidgets.QSpinBox(self.main)
        self.Holiday_Duration_input.setObjectName("Holiday_Duration_input")
        self.formLayout.setWidget(0, QtWidgets.QFormLayout.FieldRole, self.Holiday_Duration_input)
        self.Holiday_Start = QtWidgets.QLabel(self.main)
        self.Holiday_Start.setObjectName("Holiday_Start")
        self.formLayout.setWidget(1, QtWidgets.QFormLayout.LabelRole, self.Holiday_Start)
        self.Holiday_Start_input = QtWidgets.QLineEdit(self.main)
        self.Holiday_Start_input.setObjectName("Holiday_Start_input")
        self.formLayout.setWidget(1, QtWidgets.QFormLayout.FieldRole, self.Holiday_Start_input)
        self.Holiday_End = QtWidgets.QLabel(self.main)
        self.Holiday_End.setObjectName("Holiday_End")
        self.formLayout.setWidget(2, QtWidgets.QFormLayout.LabelRole, self.Holiday_End)
        self.Holiday_End_input = QtWidgets.QLineEdit(self.main)
        self.Holiday_End_input.setObjectName("Holiday_End_input")
        self.formLayout.setWidget(2, QtWidgets.QFormLayout.FieldRole, self.Holiday_End_input)
        self.comboBox = QtWidgets.QComboBox(self.main)
        self.comboBox.setObjectName("comboBox")
        self.comboBox.addItem("")
        self.comboBox.addItem("")
        self.formLayout.setWidget(3, QtWidgets.QFormLayout.FieldRole, self.comboBox)
        self.Holiday_Kind = QtWidgets.QLabel(self.main)
        self.Holiday_Kind.setObjectName("Holiday_Kind")
        self.formLayout.setWidget(3, QtWidgets.QFormLayout.LabelRole, self.Holiday_Kind)
        self.Add_Holiday_btn = QtWidgets.QPushButton(self.main)
        self.Add_Holiday_btn.setStyleSheet("background-color: rgb(255, 255, 255);")
        self.Add_Holiday_btn.setObjectName("Add_Holiday_btn")
        self.Add_Holiday_btn.clicked.connect(
            lambda: self.add_holiday(
                str(self.Military_Code.text()),
                str(self.comboBox.currentText()),
                str(self.Holiday_Duration_input.text()),
                str(self.Holiday_Start_input.text()),
                str(self.Holiday_End_input.text()),
            )
        )
        self.formLayout.setWidget(4, QtWidgets.QFormLayout.LabelRole, self.Add_Holiday_btn)
        self.gridLayout_6.addLayout(self.formLayout, 4, 1, 1, 1)
        self.Agent_Data = QtWidgets.QTextBrowser(self.main)
        self.Agent_Data.setObjectName("Agent_Data")
        self.gridLayout_6.addWidget(self.Agent_Data, 1, 1, 1, 2)
        self.Military_Code = QtWidgets.QLineEdit(self.main)
        self.Military_Code.setText("")
        self.Military_Code.setObjectName("Military_Code")
        self.gridLayout_6.addWidget(self.Military_Code, 0, 1, 1, 1)
        self.label = QtWidgets.QLabel(self.main)
        self.label.setText("")
        self.label.setObjectName("label")
        self.gridLayout_6.addWidget(self.label, 3, 1, 1, 1)
        self.Search_Bycode = QtWidgets.QPushButton(self.main)
        self.Search_Bycode.setObjectName("Search_Bycode")
        self.Search_Bycode.clicked.connect(self.Search_Btn)
        self.gridLayout_6.addWidget(self.Search_Bycode, 0, 0, 1, 1)
        self.alert_section = QtWidgets.QTableWidget(self.main)
        self.alert_section.setObjectName("alert_section")
        self.alert_section.setColumnCount(3)
        self.alert_section.setRowCount(0)
        item = QtWidgets.QTableWidgetItem()
        self.alert_section.setHorizontalHeaderItem(0, item)
        item = QtWidgets.QTableWidgetItem()
        self.alert_section.setHorizontalHeaderItem(1, item)
        item = QtWidgets.QTableWidgetItem()
        self.alert_section.setHorizontalHeaderItem(2, item)
        self.alert_section.horizontalHeader().setDefaultSectionSize(210)
        self.gridLayout_6.addWidget(self.alert_section, 8, 1, 1, 1)
        self.lineEdit = QtWidgets.QLineEdit(self.main)
        self.lineEdit.setObjectName("lineEdit")
        self.gridLayout_6.addWidget(self.lineEdit, 0, 2, 1, 1)
        self.tabWidget.addTab(self.main, "")
        self.add_agent = QtWidgets.QWidget()
        self.add_agent.setObjectName("add_agent")
        self.gridLayout_4 = QtWidgets.QGridLayout(self.add_agent)
        self.gridLayout_4.setObjectName("gridLayout_4")
        self.Civil_Registry = QtWidgets.QLabel(self.add_agent)
        self.Civil_Registry.setStyleSheet(" font-size: 15px;")
        self.Civil_Registry.setObjectName("Civil_Registry")
        self.gridLayout_4.addWidget(self.Civil_Registry, 3, 1, 1, 1)
        self.Mobile_Number = QtWidgets.QLabel(self.add_agent)
        self.Mobile_Number.setStyleSheet(" font-size: 15px;")
        self.Mobile_Number.setObjectName("Mobile_Number")
        self.gridLayout_4.addWidget(self.Mobile_Number, 4, 1, 1, 1)
        self.Governorate = QtWidgets.QLabel(self.add_agent)
        self.Governorate.setStyleSheet(" font-size: 15px;")
        self.Governorate.setObjectName("Governorate")
        self.gridLayout_4.addWidget(self.Governorate, 6, 1, 1, 1)
        self.The_Group = QtWidgets.QLabel(self.add_agent)
        self.The_Group.setStyleSheet(" font-size: 15px;")
        self.The_Group.setObjectName("The_Group")
        self.gridLayout_4.addWidget(self.The_Group, 5, 1, 1, 1)
        self.Military_Rank = QtWidgets.QLabel(self.add_agent)
        self.Military_Rank.setStyleSheet(" font-size: 15px;")
        self.Military_Rank.setObjectName("Military_Rank")
        self.gridLayout_4.addWidget(self.Military_Rank, 7, 1, 1, 1)
        self.General_Number = QtWidgets.QLabel(self.add_agent)
        self.General_Number.setStyleSheet(" font-size: 15px;")
        self.General_Number.setObjectName("General_Number")
        self.gridLayout_4.addWidget(self.General_Number, 1, 1, 1, 1)
        self.Name = QtWidgets.QLabel(self.add_agent)
        self.Name.setStyleSheet(" font-size: 15px;")
        self.Name.setObjectName("Name")
        self.gridLayout_4.addWidget(self.Name, 0, 1, 1, 1)
        self.Military_Number = QtWidgets.QLabel(self.add_agent)
        self.Military_Number.setStyleSheet(" font-size: 15px;")
        self.Military_Number.setObjectName("Military_Number")
        self.gridLayout_4.addWidget(self.Military_Number, 2, 1, 1, 1)
        self.Military_Number_input = QtWidgets.QLineEdit(self.add_agent)
        self.Military_Number_input.setText("")
        self.Military_Number_input.setObjectName("Military_Number_input")
        self.gridLayout_4.addWidget(self.Military_Number_input, 2, 2, 1, 3)
        self.agents_section = QtWidgets.QTableWidget(self.add_agent)
        self.agents_section.setObjectName("agents_section")
        self.agents_section.setColumnCount(8)
        self.agents_section.setRowCount(22)
        item = QtWidgets.QTableWidgetItem()
        self.agents_section.setVerticalHeaderItem(0, item)
        item = QtWidgets.QTableWidgetItem()
        self.agents_section.setVerticalHeaderItem(1, item)
        item = QtWidgets.QTableWidgetItem()
        self.agents_section.setVerticalHeaderItem(2, item)
        item = QtWidgets.QTableWidgetItem()
        self.agents_section.setVerticalHeaderItem(3, item)
        item = QtWidgets.QTableWidgetItem()
        self.agents_section.setVerticalHeaderItem(4, item)
        item = QtWidgets.QTableWidgetItem()
        self.agents_section.setVerticalHeaderItem(5, item)
        item = QtWidgets.QTableWidgetItem()
        self.agents_section.setVerticalHeaderItem(6, item)
        item = QtWidgets.QTableWidgetItem()
        self.agents_section.setVerticalHeaderItem(7, item)
        item = QtWidgets.QTableWidgetItem()
        self.agents_section.setVerticalHeaderItem(8, item)
        item = QtWidgets.QTableWidgetItem()
        self.agents_section.setVerticalHeaderItem(9, item)
        item = QtWidgets.QTableWidgetItem()
        self.agents_section.setVerticalHeaderItem(10, item)
        item = QtWidgets.QTableWidgetItem()
        self.agents_section.setVerticalHeaderItem(11, item)
        item = QtWidgets.QTableWidgetItem()
        self.agents_section.setVerticalHeaderItem(12, item)
        item = QtWidgets.QTableWidgetItem()
        self.agents_section.setVerticalHeaderItem(13, item)
        item = QtWidgets.QTableWidgetItem()
        self.agents_section.setVerticalHeaderItem(14, item)
        item = QtWidgets.QTableWidgetItem()
        self.agents_section.setVerticalHeaderItem(15, item)
        item = QtWidgets.QTableWidgetItem()
        self.agents_section.setVerticalHeaderItem(16, item)
        item = QtWidgets.QTableWidgetItem()
        self.agents_section.setVerticalHeaderItem(17, item)
        item = QtWidgets.QTableWidgetItem()
        self.agents_section.setVerticalHeaderItem(18, item)
        item = QtWidgets.QTableWidgetItem()
        self.agents_section.setVerticalHeaderItem(19, item)
        item = QtWidgets.QTableWidgetItem()
        self.agents_section.setVerticalHeaderItem(20, item)
        item = QtWidgets.QTableWidgetItem()
        self.agents_section.setVerticalHeaderItem(21, item)
        item = QtWidgets.QTableWidgetItem()
        self.agents_section.setHorizontalHeaderItem(0, item)
        item = QtWidgets.QTableWidgetItem()
        self.agents_section.setHorizontalHeaderItem(1, item)
        item = QtWidgets.QTableWidgetItem()
        self.agents_section.setHorizontalHeaderItem(2, item)
        item = QtWidgets.QTableWidgetItem()
        self.agents_section.setHorizontalHeaderItem(3, item)
        item = QtWidgets.QTableWidgetItem()
        self.agents_section.setHorizontalHeaderItem(4, item)
        item = QtWidgets.QTableWidgetItem()
        self.agents_section.setHorizontalHeaderItem(5, item)
        item = QtWidgets.QTableWidgetItem()
        self.agents_section.setHorizontalHeaderItem(6, item)
        item = QtWidgets.QTableWidgetItem()
        self.agents_section.setHorizontalHeaderItem(7, item)
        self.gridLayout_4.addWidget(self.agents_section, 0, 0, 10, 1)
        self.General_number_input = QtWidgets.QLineEdit(self.add_agent)
        self.General_number_input.setText("")
        self.General_number_input.setObjectName("General_number_input")
        self.gridLayout_4.addWidget(self.General_number_input, 1, 2, 1, 3)
        self.Civil_Registry_input = QtWidgets.QLineEdit(self.add_agent)
        self.Civil_Registry_input.setText("")
        self.Civil_Registry_input.setObjectName("Civil_Registry_input")
        self.gridLayout_4.addWidget(self.Civil_Registry_input, 3, 2, 1, 3)
        self.Name_input = QtWidgets.QLineEdit(self.add_agent)
        self.Name_input.setText("")
        self.Name_input.setObjectName("Name_input")
        self.gridLayout_4.addWidget(self.Name_input, 0, 2, 1, 3)
        self.The_Group_input = QtWidgets.QLineEdit(self.add_agent)
        self.The_Group_input.setText("")
        self.The_Group_input.setObjectName("The_Group_input")
        self.gridLayout_4.addWidget(self.The_Group_input, 5, 2, 1, 2)
        self.Governorate_input = QtWidgets.QLineEdit(self.add_agent)
        self.Governorate_input.setText("")
        self.Governorate_input.setObjectName("Governorate_input")
        self.gridLayout_4.addWidget(self.Governorate_input, 6, 2, 1, 1)
        self.Add_Agent_btn = QtWidgets.QPushButton(self.add_agent)
        self.Add_Agent_btn.setStyleSheet(" font-size: 15px;")
        self.Add_Agent_btn.setObjectName("Add_Agent_btn")
        self.Add_Agent_btn.clicked.connect(
            lambda: self.Add_agent(
                str(self.Military_Number_input.text()),
                str(self.Name_input.text()),
                str(self.General_number_input.text()),
                str(self.Military_Rank_input.text()),
                str(self.Governorate_input.text()),
                str(self.Mobile_Number_input.text()),
                str(self.Civil_Registry_input.text()),
                str(self.The_Group_input.text()),
            ))
        self.gridLayout_4.addWidget(self.Add_Agent_btn, 8, 2, 1, 1)
        self.Military_Rank_input = QtWidgets.QLineEdit(self.add_agent)
        self.Military_Rank_input.setText("")
        self.Military_Rank_input.setObjectName("Military_Rank_input")
        self.gridLayout_4.addWidget(self.Military_Rank_input, 7, 2, 1, 3)
        self.edit_2 = QtWidgets.QPushButton(self.add_agent)
        font = QtGui.QFont()
        font.setPointSize(12)
        self.edit_2.setFont(font)
        self.edit_2.setObjectName("edit_2")
        self.edit_2.clicked.connect(self.EditAgent)
        self.gridLayout_4.addWidget(self.edit_2, 9, 2, 1, 1)
        self.Mobile_Number_input = QtWidgets.QLineEdit(self.add_agent)
        self.Mobile_Number_input.setText("")
        self.Mobile_Number_input.setObjectName("Mobile_Number_input")
        self.gridLayout_4.addWidget(self.Mobile_Number_input, 4, 2, 1, 3)
        self.tabWidget.addTab(self.add_agent, "")
        self.gridLayout.addWidget(self.tabWidget, 0, 0, 1, 1)
        self.gridLayout_5.addLayout(self.gridLayout, 0, 0, 1, 1)
        MainWindow.setCentralWidget(self.centralwidget)
        self.menubar = QtWidgets.QMenuBar(MainWindow)
        self.menubar.setGeometry(QtCore.QRect(0, 0, 1083, 21))
        self.menubar.setLayoutDirection(QtCore.Qt.RightToLeft)
        self.menubar.setObjectName("menubar")
        self.menu = QtWidgets.QMenu(self.menubar)
        self.menu.setObjectName("menu")
        self.menu_2 = QtWidgets.QMenu(self.menubar)
        self.menu_2.setObjectName("menu_2")
        MainWindow.setMenuBar(self.menubar)
        self.statusbar = QtWidgets.QStatusBar(MainWindow)
        self.statusbar.setObjectName("statusbar")
        MainWindow.setStatusBar(self.statusbar)
        self.Add_User = QtWidgets.QAction(MainWindow)
        self.Add_User.setObjectName("Add_User")
        self.Add_User.triggered.connect(self.AdminUsersDialog)
        self.action = QtWidgets.QAction(MainWindow)
        self.action.setObjectName("action")
        self.action.triggered.connect(self.UpdateAdmin)
        self.action_2 = QtWidgets.QAction(MainWindow)
        self.action_2.setObjectName("action_2")
        self.action_2.triggered.connect(self.report)

        self.action_3 = QtWidgets.QAction(MainWindow)
        self.action_3.setObjectName("action_3")
        self.action_3.triggered.connect(self.HolidayReset)
        self.action_4 = QtWidgets.QAction(MainWindow)
        self.action_4.setObjectName("action_4")
        self.action_4.triggered.connect(self.report_bydate)
        self.menu.addSeparator()
        self.menu.addAction(self.Add_User)
        self.menu.addSeparator()
        self.menu.addAction(self.action)
        self.menu_2.addAction(self.action_2)
        self.menu_2.addSeparator()
        self.menu_2.addAction(self.action_4)
        self.menu_2.addSeparator()
        self.menu_2.addAction(self.action_3)
        self.menubar.addAction(self.menu.menuAction())
        self.menubar.addAction(self.menu_2.menuAction())

        self.retranslateUi(MainWindow)
        self.tabWidget.setCurrentIndex(0)
        QtCore.QMetaObject.connectSlotsByName(MainWindow)

    def retranslateUi(self, MainWindow):
        _translate = QtCore.QCoreApplication.translate
        MainWindow.setWindowTitle(_translate("MainWindow", "MainWindow"))
        self.Holiday_Start_input.setText(
            _translate("MainWindow", str((self.today_date())))
        )
        self.Currnet_Holiday_label.setText(_translate("MainWindow", "              الاجازة الحالية"))
        self.Current_Holiday_Kind.setText(_translate("MainWindow", "نوع الاجازة:"))
        self.Current_Holiday_Duration.setText(_translate("MainWindow", "مدة الاجازة:"))
        self.Current_Holiday_Start.setText(_translate("MainWindow", "تاريخ البداية:"))
        self.Current_Holiday_End.setText(_translate("MainWindow", "تاريخ العودة:"))
        self.Absence_Period.setText(_translate("MainWindow", "مدة التطويف :"))
        self.Remaining_Days.setText(_translate("MainWindow", "المتبقي:"))
        self.pushButton.setText(_translate("MainWindow", "تم مباشره العمل"))
        self.edit_1.setText(_translate("MainWindow", "تعديل"))
        self.edit_1.setShortcut(_translate("MainWindow", "F2"))
        self.print_1.setText(_translate("MainWindow", "طباعة"))
        self.print_1.setShortcut(_translate("MainWindow", "F1"))
        item = self.Holidays_History.horizontalHeaderItem(0)
        item.setText(_translate("MainWindow", "الرقم الخاص"))
        item = self.Holidays_History.horizontalHeaderItem(1)
        item.setText(_translate("MainWindow", "نوع الاجازة"))
        item = self.Holidays_History.horizontalHeaderItem(2)
        item.setText(_translate("MainWindow", "مدة الاجازة"))
        item = self.Holidays_History.horizontalHeaderItem(3)
        item.setText(_translate("MainWindow", "تاربخ البداية"))
        item = self.Holidays_History.horizontalHeaderItem(4)
        item.setText(_translate("MainWindow", "تاربخ النهابة"))
        item = self.Holidays_History.horizontalHeaderItem(5)
        item.setText(_translate("MainWindow", "مدة التطويف"))
        item = self.Holidays_History.horizontalHeaderItem(6)
        item.setText(_translate("MainWindow", "اسم المستخدم"))
        __sortingEnabled = self.Holidays_History.isSortingEnabled()
        self.Holidays_History.setSortingEnabled(False)
        self.Holidays_History.setSortingEnabled(__sortingEnabled)
        self.Holiday_Duration.setText(_translate("MainWindow", "مدة الاجازة :"))
        self.Holiday_Start.setText(_translate("MainWindow", "تاريخ البداية :"))
        self.Holiday_End.setText(_translate("MainWindow", "تاريخ العودة :"))
        self.comboBox.setItemText(0, _translate("MainWindow", "اعتيادية"))
        self.comboBox.setItemText(1, _translate("MainWindow", "عرضية"))
        self.Holiday_Kind.setText(_translate("MainWindow", "نوع الاجازة :"))
        self.Add_Holiday_btn.setText(_translate("MainWindow", "اضافة اجازة"))
        self.Agent_Data.setHtml(_translate("MainWindow", "<!DOCTYPE HTML PUBLIC \"-//W3C//DTD HTML 4.0//EN\" \"http://www.w3.org/TR/REC-html40/strict.dtd\">\n"
"<html><head><meta name=\"qrichtext\" content=\"1\" /><style type=\"text/css\">\n"
"p, li { white-space: pre-wrap; }\n"
"</style></head><body style=\" font-family:\'MS Shell Dlg 2\'; font-size:8.25pt; font-weight:400; font-style:normal;\">\n"
"<p style=\"-qt-paragraph-type:empty; margin-top:0px; margin-bottom:0px; margin-left:0px; margin-right:0px; -qt-block-indent:0; text-indent:0px; font-size:8pt;\"><br /></p></body></html>"))
        self.Military_Code.setPlaceholderText(_translate("MainWindow", "الرقم الخاص"))
        self.Search_Bycode.setText(_translate("MainWindow", "بحث"))
        self.Search_Bycode.setShortcut(_translate("MainWindow", "Return"))
        item = self.alert_section.horizontalHeaderItem(0)
        item.setText(_translate("MainWindow", "الرقم الخاص"))
        item = self.alert_section.horizontalHeaderItem(1)
        item.setText(_translate("MainWindow", "الاسم"))
        item = self.alert_section.horizontalHeaderItem(2)
        item.setText(_translate("MainWindow", "مدة التطويف"))
        self.lineEdit.setPlaceholderText(_translate("MainWindow", "الاسم"))
        self.tabWidget.setTabText(self.tabWidget.indexOf(self.main), _translate("MainWindow", "الرئيسية"))
        self.Civil_Registry.setText(_translate("MainWindow", "السجل المدني :"))
        self.Mobile_Number.setText(_translate("MainWindow", "رقم الجوال :"))
        self.Governorate.setText(_translate("MainWindow", "الملاك :"))
        self.The_Group.setText(_translate("MainWindow", "المجموعة :"))
        self.Military_Rank.setText(_translate("MainWindow", "الرتبة :"))
        self.General_Number.setText(_translate("MainWindow", "الرقم العام : "))
        self.Name.setText(_translate("MainWindow", "الاسم :"))
        self.Military_Number.setText(_translate("MainWindow", "الرقم الخاص :"))
        item = self.agents_section.verticalHeaderItem(0)
        item.setText(_translate("MainWindow", "1"))
        item = self.agents_section.verticalHeaderItem(1)
        item.setText(_translate("MainWindow", "2"))
        item = self.agents_section.verticalHeaderItem(2)
        item.setText(_translate("MainWindow", "4"))
        item = self.agents_section.verticalHeaderItem(3)
        item.setText(_translate("MainWindow", "5"))
        item = self.agents_section.verticalHeaderItem(4)
        item.setText(_translate("MainWindow", "6"))
        item = self.agents_section.verticalHeaderItem(5)
        item.setText(_translate("MainWindow", "7"))
        item = self.agents_section.verticalHeaderItem(6)
        item.setText(_translate("MainWindow", "8"))
        item = self.agents_section.verticalHeaderItem(7)
        item.setText(_translate("MainWindow", "9"))
        item = self.agents_section.verticalHeaderItem(8)
        item.setText(_translate("MainWindow", "10"))
        item = self.agents_section.verticalHeaderItem(9)
        item.setText(_translate("MainWindow", "11"))
        item = self.agents_section.verticalHeaderItem(10)
        item.setText(_translate("MainWindow", "12"))
        item = self.agents_section.verticalHeaderItem(11)
        item.setText(_translate("MainWindow", "13"))
        item = self.agents_section.verticalHeaderItem(12)
        item.setText(_translate("MainWindow", "14"))
        item = self.agents_section.verticalHeaderItem(13)
        item.setText(_translate("MainWindow", "15"))
        item = self.agents_section.verticalHeaderItem(14)
        item.setText(_translate("MainWindow", "16"))
        item = self.agents_section.verticalHeaderItem(15)
        item.setText(_translate("MainWindow", "17"))
        item = self.agents_section.verticalHeaderItem(16)
        item.setText(_translate("MainWindow", "18"))
        item = self.agents_section.verticalHeaderItem(17)
        item.setText(_translate("MainWindow", "19"))
        item = self.agents_section.verticalHeaderItem(18)
        item.setText(_translate("MainWindow", "20"))
        item = self.agents_section.verticalHeaderItem(19)
        item.setText(_translate("MainWindow", "21"))
        item = self.agents_section.verticalHeaderItem(20)
        item.setText(_translate("MainWindow", "22"))
        item = self.agents_section.verticalHeaderItem(21)
        item.setText(_translate("MainWindow", "23"))
        item = self.agents_section.horizontalHeaderItem(0)
        item.setText(_translate("MainWindow", "الاسم "))
        item = self.agents_section.horizontalHeaderItem(1)
        item.setText(_translate("MainWindow", "الرتية"))
        item = self.agents_section.horizontalHeaderItem(2)
        item.setText(_translate("MainWindow", "الرقم الخاص"))
        item = self.agents_section.horizontalHeaderItem(3)
        item.setText(_translate("MainWindow", "الرقم العام"))
        item = self.agents_section.horizontalHeaderItem(4)
        item.setText(_translate("MainWindow", "السجل"))
        item = self.agents_section.horizontalHeaderItem(5)
        item.setText(_translate("MainWindow", "الجوال"))
        item = self.agents_section.horizontalHeaderItem(6)
        item.setText(_translate("MainWindow", "المجموعة"))
        item = self.agents_section.horizontalHeaderItem(7)
        item.setText(_translate("MainWindow", "الملاك"))
        self.Add_Agent_btn.setText(_translate("MainWindow", "اضافة فرد"))
        self.edit_2.setText(_translate("MainWindow", "تعديل"))
        self.edit_2.setShortcut(_translate("MainWindow", "F2"))
        self.tabWidget.setTabText(self.tabWidget.indexOf(self.add_agent), _translate("MainWindow", "ادارة الافراد"))
        self.menu.setTitle(_translate("MainWindow", "مدير الوحدة"))
        self.menu_2.setTitle(_translate("MainWindow", "ادوات"))
        self.Add_User.setText(_translate("MainWindow", " اداره المستخدمين"))
        self.action.setText(_translate("MainWindow", "تعديل بيانات مدير الوحدة"))
        self.action_2.setText(_translate("MainWindow", "التقرير السنوى"))
        self.action_3.setText(_translate("MainWindow", "تحديث البيانات"))
        self.action_4.setText(_translate("MainWindow", "تقرير بالتاريخ"))